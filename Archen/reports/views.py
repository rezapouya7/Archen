# PATH: /Archen/reports/views.py

from django.shortcuts import render
from django.db.models import Count, Sum
from django.http import JsonResponse, HttpResponse
from django.contrib.auth.decorators import login_required
import os

from inventory.models import Product, Part, ProductModel, Material
from orders.models import Order
from orders.status_styles import get_status_badge_classes
from production_line.models import ProductionLog
from users.models import CustomUser
from jobs.models import ProductionJob
from utils.xlsx import base_styles, build_table_response, sanitize_value, write_table


def _xlsx_response_from_workbook(wb, filename: str) -> HttpResponse:
    """Save a workbook to an HTTP response."""
    from io import BytesIO

    bio = BytesIO()
    wb.save(bio)
    resp = HttpResponse(
        bio.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    resp['Content-Disposition'] = f"attachment; filename={filename}"
    return resp


def _prepare_detail_sheet(sheet_title: str, max_cols: int):
    """Create a RTL worksheet with shared styles and helpers."""
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title
    try:
        ws.sheet_view.rightToLeft = True
    except Exception:
        pass
    styles = base_styles()

    def set_cell(row: int, col: int, value: object, *, is_label: bool = False, alignment=None):
        c = ws.cell(row=row, column=col, value=sanitize_value(value))
        c.font = styles['header_font'] if is_label else styles['cell_font']
        c.alignment = alignment or styles['right_cell']
        c.border = styles['border']
        if is_label:
            c.fill = styles['header_fill']
        return c

    def add_title(row: int, text: str):
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=max_cols)
        c = ws.cell(row=row, column=1, value=text)
        c.font = styles['title_font']
        c.alignment = styles['center_header']
        return c

    return wb, ws, styles, set_cell, add_title

def _pdf_response_from_html(html: str, filename: str, request=None) -> HttpResponse:
    """Render a PDF using ReportLab only (Persian-safe).

    Notes:
    - Shapes Persian text using arabic_reshaper + python-bidi so the
      final PDF shows correct glyph order.
    - Registers Vazirmatn TTF from static if available.
    - Performs a very light HTML-to-text conversion, preserving line
      breaks and adding simple separators for table cells/rows.
    """
    from io import BytesIO
    from django.conf import settings
    from reportlab.lib.pagesizes import A4  # type: ignore
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer  # type: ignore
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle  # type: ignore
    from reportlab.lib.enums import TA_RIGHT  # type: ignore
    from reportlab.pdfbase import pdfmetrics  # type: ignore
    from reportlab.pdfbase.ttfonts import TTFont  # type: ignore

    # Register Persian font (Vazirmatn) if available under static
    def _find_font(relpath: str) -> str | None:
        sr = getattr(settings, 'STATIC_ROOT', None)
        if sr:
            cand = os.path.join(str(sr), *(relpath.split('/')))
            if os.path.exists(cand):
                return cand
        try:
            from django.contrib.staticfiles import finders  # type: ignore
            found = finders.find(relpath)
            if found and os.path.exists(found):
                return found
        except Exception:
            return None
        # Fallback: try project static directory directly
        try:
            base = getattr(settings, 'BASE_DIR', None)
            if base:
                cand = os.path.join(str(base), 'static', *relpath.split('/'))
                if os.path.exists(cand):
                    return cand
        except Exception:
            pass
        return None

    bold_name = None
    try:
        reg = _find_font('fonts/Vazirmatn/Vazirmatn-Regular.ttf')
        bold = _find_font('fonts/Vazirmatn/Vazirmatn-Bold.ttf')
        if reg and os.path.exists(reg):
            pdfmetrics.registerFont(TTFont('Vazirmatn', reg))
            font_name = 'Vazirmatn'
            if bold and os.path.exists(bold):
                pdfmetrics.registerFont(TTFont('Vazirmatn-Bold', bold))
                bold_name = 'Vazirmatn-Bold'
        else:
            font_name = 'Helvetica'
    except Exception:
        font_name = 'Helvetica'

    # Persian shaping (visual order for ReportLab)
    try:
        from arabic_reshaper import reshape  # type: ignore
        from bidi.algorithm import get_display  # type: ignore
        import re as _re

        def fa(s: str) -> str:
            """Apply Arabic reshaper + bidi with explicit RTL base direction.

            - If the string has Persian/Arabic letters, return shaped visual order
              so ReportLab shows connected glyphs in right-to-left.
            - If it is purely numbers/date/latin, return as-is to preserve order.
            """
            try:
                if not s:
                    return s
                if _re.search(r"[\u0600-\u06FF]", s):
                    return get_display(reshape(s), base_dir='R')
                return s
            except Exception:
                return s
    except Exception:
        def fa(s: str) -> str:
            return s

    # Minimal HTML renderer: remove style/script, render paragraphs and tables
    import re, html as ihtml

    # Drop entire <head> (contains style/script) and any stray style/script/comments anywhere
    safe_html = re.sub(r'<\s*head[^>]*>.*?<\s*/\s*head\s*>', '', html, flags=re.I | re.S)
    safe_html = re.sub(r'<\s*style\b[^>]*>.*?<\s*/\s*style\s*>', '', safe_html, flags=re.I | re.S)
    safe_html = re.sub(r'<\s*script\b[^>]*>.*?<\s*/\s*script\s*>', '', safe_html, flags=re.I | re.S)
    safe_html = re.sub(r'<!--.*?-->', '', safe_html, flags=re.S)

    buf = BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=A4, rightMargin=36, leftMargin=36, topMargin=36, bottomMargin=36
    )
    styles = getSampleStyleSheet()
    p_style = ParagraphStyle(
        'fa', parent=styles['Normal'], fontName=font_name, fontSize=11, leading=16, alignment=TA_RIGHT
    )

    from reportlab.platypus import Table, LongTable, TableStyle  # type: ignore
    from reportlab.lib import colors  # type: ignore

    def strip_tags(s: str) -> str:
        s = re.sub(r'<[^>]+>', '', s)
        s = ihtml.unescape(s)
        return s.strip()

    def render_paragraph_block(fragment: str, story: list) -> None:
        frag = fragment
        frag = re.sub(r'<\s*br\s*/?>', '\n', frag, flags=re.I)
        # Replace block tags with newlines to split into logical lines
        frag = re.sub(r'</\s*(p|div|h1|h2|h3|h4|h5|h6)\s*>', '\n', frag, flags=re.I)
        text = strip_tags(frag)
        # Split to lines; add paragraphs
        for ln in [x for x in text.splitlines() if x.strip()]:
            story.append(Paragraph(fa(ln.strip()), p_style))
            story.append(Spacer(1, 6))

    def render_table_block(table_html: str, story: list) -> None:
        # Extract rows
        rows = []
        has_header = False
        for r in re.findall(r'<\s*tr[^>]*>(.*?)</\s*tr\s*>', table_html, flags=re.I | re.S):
            # Capture both th and td in order
            cells = re.findall(r'<\s*(t[dh])[^>]*>(.*?)</\s*t[dh]\s*>', r, flags=re.I | re.S)
            if not cells:
                continue
            row = [fa(strip_tags(c_html)) for tag, c_html in cells]
            rows.append(row)
            if any(tag.lower() == 'th' for tag, _ in cells):
                has_header = True
        if not rows:
            return
        # Normalize row lengths
        maxc = max(len(r) for r in rows)
        rows = [r + [''] * (maxc - len(r)) for r in rows]

        # Use LongTable to ensure multi-page tables render fully
        try:
            tbl = LongTable(rows, hAlign='RIGHT', repeatRows=1 if has_header else 0)
        except Exception:
            tbl = Table(rows, hAlign='RIGHT', repeatRows=1 if has_header else 0)
        ts = [
            ('FONTNAME', (0, 0), (-1, -1), font_name),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('ALIGN', (0, 0), (-1, -1), 'RIGHT'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            # Darker grid to match list tables better
            ('GRID', (0, 0), (-1, -1), 0.8, colors.HexColor('#D1D5DB')),
            ('LEFTPADDING', (0, 0), (-1, -1), 4),
            ('RIGHTPADDING', (0, 0), (-1, -1), 4),
            ('TOPPADDING', (0, 0), (-1, -1), 3),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
        ]
        if has_header:
            ts += [
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#F9FAFB')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor('#111827')),
            ]
            if bold_name:
                ts.append(('FONTNAME', (0, 0), (-1, 0), bold_name))
        tbl.setStyle(TableStyle(ts))
        story.append(tbl)
        story.append(Spacer(1, 8))

    story = []
    # Iterate content: split around <table> blocks, render text between
    pos = 0
    for m in re.finditer(r'<\s*table[^>]*>.*?</\s*table\s*>', safe_html, flags=re.I | re.S):
        before = safe_html[pos:m.start()]
        if before.strip():
            render_paragraph_block(before, story)
        render_table_block(m.group(0), story)
        pos = m.end()
    # Remainder after last table
    tail = safe_html[pos:]
    if tail.strip():
        render_paragraph_block(tail, story)

    # Build PDF
    if not story:
        # Fallback: render plain stripped text to avoid empty PDF
        render_paragraph_block(safe_html, story)
    doc.build(story)
    pdf = buf.getvalue()
    resp = HttpResponse(pdf, content_type='application/pdf')
    resp['Content-Disposition'] = f"attachment; filename={filename}.pdf"
    return resp
    doc.build(story)
    pdf = buf.getvalue()
    resp = HttpResponse(pdf, content_type='application/pdf')
    resp['Content-Disposition'] = f"attachment; filename={filename}.pdf"
    return resp
def _gather_reports_metrics():
    """
    Compute and return all datasets used by the reports dashboard.

    This central helper is used by both the HTML view and the JSON API to keep
    logic consistent. It helps the frontend refresh data without a full page
    reload and stay in sync with inventory changes.
    """
    total_products = Product.objects.count()
    total_parts = Part.objects.count()
    total_orders = Order.objects.count()
    total_production_logs = ProductionLog.objects.count()
    # Additional totals for extra cards
    total_models = ProductModel.objects.count()
    total_materials = Material.objects.count()
    total_users = CustomUser.objects.count()
    total_jobs = ProductionJob.objects.count()

    # Count orders by status for the status chart (used both in static and dynamic views)
    all_statuses = [choice[0] for choice in Order.STATUS_CHOICES]
    status_counts_map = {label: 0 for label in all_statuses}
    status_counts = list(
        Order.objects.values('status').annotate(count=Count('id')).order_by('status')
    )
    for item in status_counts:
        label = item.get('status') or ''
        if label in status_counts_map:
            status_counts_map[label] = item['count']
        else:
            status_counts_map[label] = item['count']
    chart_labels = list(status_counts_map.keys())
    chart_data = [status_counts_map[label] for label in chart_labels]
    # Colors aligned with the Orders app badge backgrounds (orders_list.html)
    orders_status_color_map = {
        'در انتظار': '#e5e7eb',   # bg-gray-200
        'در حال ساخت': '#fde68a', # bg-amber-200
        'در انبار': '#bfdbfe',     # bg-blue-200
        'ارسال شده': '#bbf7d0',    # bg-green-200
        'لغو شده': '#fecaca',      # bg-red-200
        'گارانتی': '#99f6e4',      # bg-teal-200
    }
    chart_colors = [orders_status_color_map.get(label, '#93c5fd') for label in chart_labels]
    orders_status_class_map = {
        label: get_status_badge_classes(label)
        for label in chart_labels
    }
    orders_status_summary = [
        {
            'label': label,
            'count': status_counts_map.get(label, 0),
            'classes': orders_status_class_map.get(label, 'bg-gray-200 text-gray-800'),
        }
        for label in chart_labels
    ]

    # Additional datasets for interactive charts
    # Products by product model name
    products_by_type = Product.objects.values('product_model__name').annotate(count=Count('id')).order_by('product_model__name')
    products_chart_labels = []
    products_chart_data = []
    products_summary = []
    for item in products_by_type:
        key = item.get('product_model__name') or ''
        cnt = item['count']
        products_chart_labels.append(key)
        products_chart_data.append(cnt)
        products_summary.append({'label': key, 'count': cnt})

    # Parts by product model name
    parts_by_type = Part.objects.values('product_model__name').annotate(count=Count('id')).order_by('product_model__name')
    parts_chart_labels = []
    parts_chart_data = []
    parts_summary = []
    for item in parts_by_type:
        key = item.get('product_model__name') or ''
        cnt = item['count']
        parts_chart_labels.append(key)
        parts_chart_data.append(cnt)
        parts_summary.append({'label': key, 'count': cnt})

    # Production logs and open jobs by product-based sections (7 buckets)
    # English: The dashboard chart for logs should visualise how many
    # jobs are registered vs still open in each production section.
    # We align these buckets with the product stock sections (assembly
    # through packaging) and expose both series so the frontend can
    # render grouped columns.
    from production_line.models import SectionChoices
    product_sections_order = [
        SectionChoices.ASSEMBLY,
        SectionChoices.WORKPAGE,
        SectionChoices.UNDERCOATING,
        SectionChoices.PAINTING,
        SectionChoices.SEWING,
        SectionChoices.UPHOLSTERY,
        SectionChoices.PACKAGING,
    ]
    section_label_map = dict(SectionChoices.choices)

    # Human‑readable labels (Persian) for the x‑axis
    logs_chart_labels = [section_label_map.get(code, code) for code in product_sections_order]

    # Registered work: count ProductionLog rows per section
    logs_raw = (
        ProductionLog.objects
        .filter(section__in=product_sections_order)
        .values('section')
        .annotate(count=Count('id'))
    )
    registered_map = {code: 0 for code in product_sections_order}
    for item in logs_raw:
        code = item.get('section')
        if code in registered_map:
            registered_map[code] = item['count']

    logs_chart_registered_data = [registered_map[code] for code in product_sections_order]

    # Open work: count jobs that appear in the daily work-entry dropdown
    # for each section.  Mirror the logic from ``production_line.views.work_entry``
    # so that the chart matches the "شماره کار" list shown to operators.
    open_map = {code: 0 for code in product_sections_order}
    logs_open_jobs_list = []
    # Consider only unfinished jobs
    jobs_qs = (
        ProductionJob.objects
        .filter(finished_at__isnull=True)
        .select_related('product__product_model', 'part')
    )
    job_ids = list(jobs_qs.values_list('id', flat=True))
    job_logs = {}
    if job_ids:
        for row in (
            ProductionLog.objects
            .filter(job_id__in=job_ids, section__in=product_sections_order)
            .values('job_id', 'section')
        ):
            job_logs.setdefault(row['job_id'], set()).add(row['section'])
    # Normalise allowed sections order as in work_entry view
    ORDER = ['assembly', 'workpage', 'undercoating', 'painting', 'sewing', 'upholstery', 'packaging']
    label_display_map = dict(ProductionJob.LABEL_CHOICES)
    for job in jobs_qs:
        logs_for_job = job_logs.get(job.id, set())
        allowed = list(getattr(job, 'allowed_sections', []) or [])
        allowed_norm = []
        if allowed:
            allowed_lower = {str(x).lower() for x in allowed}
            allowed_norm = [s for s in ORDER if s in allowed_lower]
        for section_code in product_sections_order:
            # Ensure we compare using the plain slug string
            section_slug = str(section_code)
            # Equivalent to ``exclude(productionlog__section=section_slug)``
            if section_slug in logs_for_job:
                continue
            if allowed_norm:
                if section_slug not in allowed_norm:
                    continue
                idx = allowed_norm.index(section_slug)
                if idx > 0:
                    prev = allowed_norm[idx - 1]
                    # Only open when previous section has at least one log
                    if prev not in logs_for_job:
                        continue
            open_map[section_code] += 1
            # Build an entry for the open jobs list (used in logs panel)
            product = getattr(job, 'product', None)
            part = getattr(job, 'part', None)
            model_name = ''
            if product is not None:
                model_name = getattr(getattr(product, 'product_model', None), 'name', '') or getattr(product, 'name', '')
            elif part is not None:
                model_name = getattr(part, 'product_model', None) or ''
            item_name = ''
            if part is not None:
                item_name = getattr(part, 'name', '')
            elif product is not None:
                item_name = getattr(product, 'name', '')
            # English: For the open jobs list, compute Jalali creation date/time
            # strings similar to the logs table jdate + time columns.
            try:
                g_created = timezone.localtime(job.created_at) if job.created_at else None
                if g_created is not None:
                    j_created = jdatetime.datetime.fromgregorian(datetime=g_created)
                    created_date = j_created.strftime('%Y-%m-%d')
                    created_time = j_created.strftime('%H:%M')
                else:
                    created_date = ''
                    created_time = ''
            except Exception:
                if job.created_at:
                    created_date = job.created_at.strftime('%Y-%m-%d')
                    created_time = job.created_at.strftime('%H:%M')
                else:
                    created_date = ''
                    created_time = ''

            logs_open_jobs_list.append(
                {
                    'job_number': job.job_number,
                    'section': section_slug,
                    'section_label': section_label_map.get(section_code, section_slug),
                    'model': model_name,
                    'item_name': item_name,
                    'label_display': label_display_map.get(job.job_label, ''),
                    'created_date': created_date,
                    'created_time': created_time,
                }
            )

    logs_chart_open_data = [open_map[code] for code in product_sections_order]

    # Preserve a single series for any legacy usage (registered count)
    logs_chart_data = list(logs_chart_registered_data)

    logs_summary = [
        {
            'label': section_label_map.get(code, code),
            'registered': registered_map.get(code, 0),
            'open': open_map.get(code, 0),
        }
        for code in product_sections_order
	    ]

    # Models card should show distribution of PARTS per model (not products)
    models_chart_labels = list(parts_chart_labels)
    models_chart_data = list(parts_chart_data)
    models_summary = list(parts_summary)

    # Materials dataset: shortage (<= threshold) vs normal counts + list of low-stock materials
    below = 0
    normal = 0
    low_stock_list = []
    for m in Material.objects.all():
        qty = m.quantity or 0
        thr = m.threshold or 0
        diff = qty - thr
        # English: Treat materials with quantity less than OR equal to threshold as shortage
        if qty <= thr:
            below += 1
            low_stock_list.append({'label': m.name, 'count': qty})
        else:
            normal += 1
    # Sort low stock ascending by quantity and take top 10
    low_stock_list.sort(key=lambda x: (x['count'] if x['count'] is not None else 0))
    # Persian labels for the materials inventory chart
    # English: Replace the below-threshold label with a clearer shortage warning
    materials_chart_labels = ['کمبود موجودی', 'نرمال']
    materials_chart_data = [below, normal]
    materials_summary = low_stock_list[:10]

    # Parts inventory dataset: count parts by shortage status.
    # English comment: A part is in shortage if ANY unit bucket (cut or cnc/tools)
    # is at or below threshold (<=). Normal only when BOTH buckets are above.
    below_p = 0
    normal_p = 0
    low_parts = []
    for p in Part.objects.all():
        cut = getattr(p, 'stock_cut', 0) or 0
        cnc = getattr(p, 'stock_cnc_tools', 0) or 0
        thr = getattr(p, 'threshold', 0) or 0
        is_low = (cut <= thr) or (cnc <= thr)
        # For summary sorting, use the lower unit bucket as the key
        min_unit_stock = cut if cut <= cnc else cnc
        if is_low:
            below_p += 1
            low_parts.append({'label': p.name, 'count': min_unit_stock})
        else:
            normal_p += 1
    low_parts.sort(key=lambda x: (x['count'] if x['count'] is not None else 0))
    # Use simplified labels without parenthetical threshold hints for display
    parts_inventory_chart_labels = ['کمبود موجودی', 'نرمال']
    parts_inventory_chart_data = [below_p, normal_p]
    parts_inventory_summary = low_parts[:10]

    # Users dataset: counts by role
    role_map = dict(CustomUser.ROLE_CHOICES)
    users_by_role = CustomUser.objects.values('role').annotate(count=Count('id')).order_by('role')
    users_chart_labels = []
    users_chart_data = []
    users_summary = []
    users_chart_colors = []
    # Colors aligned with Users app badges (user_list.html)
    role_color_map = {
        'manager': '#2563eb',            # blue-600
        'accountant': '#d97706',         # amber-600
        'seller': '#9333ea',             # purple-600
        'cutter_master': '#0284c7',      # sky-600
        'cnc_master': '#4f46e5',         # indigo-600
        'undercoating_master': '#a16207',# yellow-700
        'painting_master': '#e11d48',    # rose-600
        'assembly_master': '#0d9488',    # teal-600
        'sewing_master': '#c026d3',      # fuchsia-600
        'upholstery_master': '#16a34a',  # green-600
        'packaging_master': '#ea580c',   # orange-600
        'workpage_master': '#65a30d',    # lime-600
    }
    for item in users_by_role:
        code = item['role']
        label = role_map.get(code, code)
        cnt = item['count']
        users_chart_labels.append(label)
        users_chart_data.append(cnt)
        users_summary.append({'label': label, 'count': cnt})
        users_chart_colors.append(role_color_map.get(code, '#4b5563'))  # gray-600 default

    # Jobs dataset: counts by job_label (with consistent colors matching job list badges)
    label_map = dict(ProductionJob.LABEL_CHOICES)
    label_order = [code for code, _ in ProductionJob.LABEL_CHOICES]
    label_counts = {code: 0 for code in label_order}
    jobs_by_label = ProductionJob.objects.values('job_label').annotate(count=Count('id')).order_by('job_label')
    jobs_chart_labels = []
    jobs_chart_data = []
    jobs_chart_colors = []
    jobs_summary = []
    jobs_status_summary = []
    # Color map based on badges used in jobs app and production_line work entry
    label_colors = {
        'in_progress': '#6b7280',  # gray-500
        'completed':   '#68d391',  # green-400
        'scrapped':    '#dc2626',  # red-600
        'warranty':    '#fcd34d',  # yellow-300
        'repaired':    '#2563eb',  # blue-600
        'deposit':     '#8B4513',  # brown
    }
    job_badge_classes = {
        'in_progress': 'bg-gray-500 text-white',
        'completed': 'bg-green-400 text-white',
        'scrapped': 'bg-red-600 text-white',
        'warranty': 'bg-yellow-300 text-black',
        'repaired': 'bg-blue-600 text-white',
        'deposit': 'text-white',
    }
    job_badge_styles = {
        'deposit': 'background-color:#8B4513',
    }
    for item in jobs_by_label:
        code = item.get('job_label') or ''
        if code in label_counts:
            label_counts[code] = item['count']
        else:
            label_counts[code] = item['count']
            label_order.append(code)
    for code in label_order:
        label = label_map.get(code, code)
        cnt = label_counts.get(code, 0)
        color = label_colors.get(code, '#6b7280')
        jobs_chart_labels.append(label)
        jobs_chart_data.append(cnt)
        jobs_chart_colors.append(color)
        jobs_summary.append({'label': label, 'count': cnt})
        jobs_status_summary.append({
            'code': code,
            'label': label,
            'count': cnt,
            'classes': job_badge_classes.get(code, 'bg-gray-400 text-white'),
            'style': job_badge_styles.get(code, ''),
            'color': color,
        })

    # Jobs list for Job Details panel (same dataset as main jobs list)
    # English: Use the full queryset ordered by creation time so that the
    # dashboard list stays in sync with the Jobs module list view.
    jobs_list_qs = (
        ProductionJob.objects
        .select_related('product__product_model', 'part')
        .order_by('-created_at')
    )

    # Lightweight orders list for Order Details panel (recent 200)
    orders_list_qs = (
        Order.objects
        .order_by('-id')[:200]
    )

    # Logs list for Log Details panel (show all records without limit)
    logs_list_qs = (
        ProductionLog.objects
        .select_related('job', 'user', 'product', 'part')
        .order_by('-logged_at', '-id')
    )

    # English: Provide section choices in a specific UI order for the logs filter dropdown
    from production_line.models import SectionChoices
    preferred_order = [
        SectionChoices.CUTTING,
        SectionChoices.CNC_TOOLS,
        SectionChoices.ASSEMBLY,
        SectionChoices.WORKPAGE,
        SectionChoices.UNDERCOATING,
        SectionChoices.PAINTING,
        SectionChoices.SEWING,
        SectionChoices.UPHOLSTERY,
        SectionChoices.PACKAGING,
    ]
    section_label_map = dict(SectionChoices.choices)
    section_choices_ordered = [(code, section_label_map.get(code, code)) for code in preferred_order]

    # English: Provide additional choices for filters: users and product models
    user_choices = []
    for u in CustomUser.objects.all().order_by('full_name', 'username'):
        label = getattr(u, 'full_name', None) or getattr(u, 'username', '')
        user_choices.append((str(u.id), label))
    model_choices = list(ProductModel.objects.values_list('name', flat=True).order_by('name'))

    context = {
        'total_products': total_products,
        'total_parts': total_parts,
        'total_orders': total_orders,
        'total_production_logs': total_production_logs,
        'total_models': total_models,
        'total_materials': total_materials,
        'total_users': total_users,
        'total_jobs': total_jobs,
        # Static orders status chart
        'chart_labels': chart_labels,
        'chart_data': chart_data,
        'chart_colors': chart_colors,
        'orders_status_summary': orders_status_summary,
        'orders_status_class_map': orders_status_class_map,
        # Datasets for dynamic charts
        'products_chart_labels': products_chart_labels,
        'products_chart_data': products_chart_data,
        'parts_chart_labels': parts_chart_labels,
        'parts_chart_data': parts_chart_data,
        'logs_chart_labels': logs_chart_labels,
        'logs_chart_data': logs_chart_data,
        'logs_chart_registered_data': logs_chart_registered_data,
        'logs_chart_open_data': logs_chart_open_data,
        # Additional datasets for new cards
        'models_chart_labels': models_chart_labels,
        'models_chart_data': models_chart_data,
        'parts_inventory_chart_labels': parts_inventory_chart_labels,
        'parts_inventory_chart_data': parts_inventory_chart_data,
        'materials_chart_labels': materials_chart_labels,
        'materials_chart_data': materials_chart_data,
        'users_chart_labels': users_chart_labels,
        'users_chart_data': users_chart_data,
        'users_chart_colors': users_chart_colors,
        'jobs_chart_labels': jobs_chart_labels,
        'jobs_chart_data': jobs_chart_data,
        'jobs_chart_colors': jobs_chart_colors,
        # Summaries for list rendering
        'products_summary': products_summary,
        'parts_summary': parts_summary,
	        'logs_summary': logs_summary,
	        'logs_open_jobs_list': logs_open_jobs_list,
        'models_summary': models_summary,
        'materials_summary': materials_summary,
        'parts_inventory_summary': parts_inventory_summary,
        'users_summary': users_summary,
        'jobs_summary': jobs_summary,
        'jobs_status_summary': jobs_status_summary,
        'jobs_list': jobs_list_qs,
        'orders_list': orders_list_qs,
        'logs_list': logs_list_qs,
        # Section choices for UI filters (logs panel) in preferred order
        'section_choices': section_choices_ordered,
        # Extra filters: users and models
        'user_choices': user_choices,
        'model_choices': model_choices,
    }
    return context


@login_required(login_url="/users/login/")
def index(request):
    """Render the reports dashboard page using the computed metrics."""
    context = _gather_reports_metrics()
    return render(request, 'reports/index.html', context)


@login_required(login_url="/users/login/")
def metrics_api(request):
    """
    Lightweight JSON endpoint for live dashboard refresh.

    Frontend polls this endpoint to update counters and charts after DB changes,
    fixing out-of-sync issues such as the parts inventory status card.
    """
    ctx = _gather_reports_metrics()
    data = {
        'totals': {
            'products': ctx['total_products'],
            'parts': ctx['total_parts'],
            'orders': ctx['total_orders'],
            'logs': ctx['total_production_logs'],
            'models': ctx['total_models'],
            'materials': ctx['total_materials'],
            'users': ctx['total_users'],
            'jobs': ctx['total_jobs'],
        },
        'datasets': {
            'products': {
                'labels': ctx['products_chart_labels'],
                'data': ctx['products_chart_data'],
            },
            'parts': {
                'labels': ctx['parts_inventory_chart_labels'],
                'data': ctx['parts_inventory_chart_data'],
            },
            'orders': {
                'labels': ctx['chart_labels'],
                'data': ctx['chart_data'],
                'backgroundColor': ctx['chart_colors'],
            },
            'logs': {
                'labels': ctx['logs_chart_labels'],
                'data': ctx['logs_chart_data'],
                'series': [
                    {
                        'key': 'registered',
                        'label': 'ثبت شده',
                        'data': ctx.get('logs_chart_registered_data', []),
                        'backgroundColor': '#4ade80',
                    },
                    {
                        'key': 'open',
                        'label': 'باز',
                        'data': ctx.get('logs_chart_open_data', []),
                        'backgroundColor': '#f97316',
                    },
                ],
            },
            'models': {
                'labels': ctx['models_chart_labels'],
                'data': ctx['models_chart_data'],
            },
            'materials': {
                'labels': ctx['materials_chart_labels'],
                'data': ctx['materials_chart_data'],
            },
            'users': {
                'labels': ctx['users_chart_labels'],
                'data': ctx['users_chart_data'],
                'backgroundColor': ctx['users_chart_colors'],
            },
            'jobs': {
                'labels': ctx['jobs_chart_labels'],
                'data': ctx['jobs_chart_data'],
                'backgroundColor': ctx['jobs_chart_colors'],
            },
        },
    }
    return JsonResponse(data)

@login_required(login_url="/users/login/")
def scrap_report(request):
    """
    Aggregate scrap ProductionLog entries with filters by date range and section/model.
    """
    from django.db.models import Count
    from production_line.models import SectionChoices
    qs = ProductionLog.objects.filter(is_scrap=True)

    # Filters
    section = request.GET.get('section') or ''
    model = request.GET.get('model') or ''
    date_from = request.GET.get('date_from') or ''
    date_to = request.GET.get('date_to') or ''

    if section:
        qs = qs.filter(section=section)
    if model:
        qs = qs.filter(model__iexact=model)
    if date_from:
        qs = qs.filter(jdate__gte=date_from)
    if date_to:
        qs = qs.filter(jdate__lte=date_to)

    # Group by section and model
    agg = qs.values('section', 'model').annotate(count=Count('id')).order_by('section','model')
    data = []
    for row in agg:
        data.append({
            'section': dict(SectionChoices.choices).get(row['section'], row['section']),
            'model': row['model'] or '',
            'count': row['count'],
        })

    # Distinct lists for filters
    sections = list(SectionChoices.choices)
    models = list(ProductionLog.objects.exclude(model__isnull=True).exclude(model__exact='').values_list('model', flat=True).distinct().order_by('model'))

    return render(request, 'reports/scrap_report.html', {
        'data': data,
        'sections': sections,
        'models': models,
        'filters': {'section': section, 'model': model, 'date_from': date_from, 'date_to': date_to},
    })


from django.core.paginator import Paginator
# Import the relocated ProductionJob model from the jobs app.  SectionChoices
# and ProductionLog remain in the production_line app.
from jobs.models import ProductionJob
from production_line.models import ProductionLog, SectionChoices

@login_required(login_url="/users/login/")
def jobs(request):
    """List Production Jobs with colored badges and basic filters.

    The ``status`` query-string parameter is accepted in two forms for
    backward compatibility:

    - Internal code (e.g. ``completed``, ``in_progress``) matching
      ``ProductionJob.job_label`` values.
    - Persian display label (e.g. ``تولید شده``) that may be generated
      by older links or UI elements.

    When a Persian label is received it is mapped back to the internal
    code before filtering so that legacy links continue to work and the
    dropdown filter remains in sync with the actual query.
    """
    # Prefetch related order and its items/products to display order details per job
    qs = (
        ProductionJob.objects
        .select_related('product', 'part', 'order')
        .prefetch_related('order__items__product')
        .order_by('-id')
    )

    # Read filter from query string and normalise it to an internal code
    raw_status = (request.GET.get('status') or '').strip()
    status_code = raw_status
    if raw_status:
        # Map of internal code -> Persian label
        label_map = dict(ProductionJob.LABEL_CHOICES)
        if status_code not in label_map:
            # Attempt to resolve using the Persian label coming from old links
            reverse_map = {v: k for k, v in label_map.items()}
            status_code = reverse_map.get(raw_status, '')
        if status_code:
            qs = qs.filter(job_label=status_code)

    paginator = Paginator(qs, 25)
    page_obj = paginator.get_page(request.GET.get('page'))

    # Use the same label colors used in the production line app to keep UI consistent
    label_colors = {
        'in_progress': '#6b7280',  # gray-500
        'completed':   '#68d391',  # green-400
        'scrapped':    '#dc2626',  # red-600
        'warranty':    '#fcd34d',  # yellow-300
        'repaired':    '#2563eb',  # blue-600
        'deposit':     '#8B4513',  # brown
    }
    label_text_colors = {
        'in_progress': '#ffffff',
        'completed':   '#000000',
        'scrapped':    '#ffffff',
        'warranty':    '#000000',
        'repaired':    '#ffffff',
        'deposit':     '#ffffff',
    }

    # Attach computed colors to each job in the current page for easy templating
    for j in page_obj.object_list:
        jl = getattr(j, 'job_label', '') or 'in_progress'
        setattr(j, '_label_color', label_colors.get(jl, '#6b7280'))
        setattr(j, '_label_text_color', label_text_colors.get(jl, '#ffffff'))

    context = {
        'page_obj': page_obj,
        # Expose the normalised internal code so the dropdown stays in sync
        'status': status_code,
        'status_choices': ProductionJob.LABEL_CHOICES,
        'label_colors': label_colors,
        'label_text_colors': label_text_colors,
    }
    return render(request, 'reports/jobs.html', context)


@login_required(login_url="/users/login/")
def job_detail(request, job_number: str):
    """
    Display detailed information about a single production_line job.

    Shows the product name and model, a chronological list of production_line
    log entries (with user, date and time, and flags), and the total
    duration of the job from the first to the last log.  If the job is
    not found a 404 is raised.  The view does not require manager role
    and can be accessed by any authenticated user.
    """
    from django.shortcuts import get_object_or_404
    job = get_object_or_404(ProductionJob, job_number=job_number)
    # Gather logs related to this job in chronological order
    logs = (
        ProductionLog.objects
        .filter(job=job)
        .select_related('user')
        .order_by('logged_at', 'id')
    )
    log_entries = []
    first_time = None
    last_time = None
    for log in logs:
        dt = getattr(log, 'logged_at', None)
        if dt and not first_time:
            first_time = dt
        if dt:
            last_time = dt
        # Determine human friendly section name
        section_label = dict(SectionChoices.choices).get(log.section, log.section)
        log_entries.append({
            'jdate': getattr(log, 'jdate', ''),
            'time': dt.strftime('%H:%M') if dt else '-',
            'section': section_label,
            'user': (getattr(log.user, 'full_name', None) or getattr(log.user, 'username', '')),
            'is_scrap': log.is_scrap,
            'is_external': log.is_external,
            'note': log.note or '',
        })
    # Compute duration
    total_duration = None
    if first_time and last_time:
        delta = last_time - first_time
        # Format as hours:minutes
        total_duration = f"{delta.days * 24 + delta.seconds // 3600:02d}:{(delta.seconds % 3600) // 60:02d}"
    context = {
        'job': job,
        'log_entries': log_entries,
        'total_duration': total_duration,
    }
    return render(request, 'reports/job_detail.html', context)


# ------------------------------------------------------------------
# Jobs report: list and detail
# ------------------------------------------------------------------
from django.shortcuts import get_object_or_404
from django.utils.timezone import now
from jobs.models import ProductionJob

@login_required(login_url="/users/login/")
def jobs_list(request):
    jobs = ProductionJob.objects.all().order_by('-created_at')
    context = {'jobs': jobs}
    return render(request, 'reports/jobs_list.html', context)

@login_required(login_url="/users/login/")
def job_detail(request, job_id: int):
    job = get_object_or_404(ProductionJob, pk=job_id)
    logs = ProductionLog.objects.filter(job=job).order_by('logged_at')
    # Compute elapsed production time (first to last log)
    elapsed = None
    if logs:
        elapsed = (logs.last().logged_at - logs.first().logged_at)
    context = {'job': job, 'logs': logs, 'elapsed': elapsed}
    return render(request, 'reports/job_detail.html', context)

# -------------------------------------------------------------
# Job Details panel (dashboard) + Export endpoints
# -------------------------------------------------------------
from django.http import JsonResponse, HttpResponse, HttpResponseBadRequest
from django.template.loader import render_to_string
from django.utils.encoding import smart_str
from inventory.models import ProductMaterial, ProductComponent
from django.utils import timezone
import jdatetime

def _compute_job_consumption(job: ProductionJob) -> dict:
    """Compute actual internal consumption for a job based on assembly logs.

    - Count assembly logs for this job that are not external; both produced and scrap
      entries consume one unit of BOM.
    - Multiply ProductComponent and ProductMaterial quantities by the unit count.
    This is an approximation aligned with current logging (+1 per log).
    """
    if not job.product_id:
        return {"parts": [], "materials": [], "units": 0}
    base_qs = ProductionLog.objects.filter(job=job, section='assembly', is_external=False)
    units = base_qs.count()  # each log = 1 unit (produced or scrap)

    parts = []
    for row in ProductComponent.objects.filter(product=job.product).select_related('part'):
        qty = int(row.qty or 0) * units
        if qty > 0 and row.part_id:
            parts.append({
                'name': getattr(row.part, 'name', ''),
                'qty': qty,
            })
    materials = []
    for row in ProductMaterial.objects.filter(product=job.product).select_related('material'):
        try:
            req = float(row.qty)
        except Exception:
            req = 0.0
        qty = req * units
        if qty > 0 and row.material_id:
            materials.append({
                'name': getattr(row.material, 'name', ''),
                'unit': getattr(row.material, 'unit', ''),
                'qty': qty,
            })
    return {"parts": parts, "materials": materials, "units": units}


@login_required(login_url="/users/login/")
def job_details_panel(request):
    """Return HTML panel for a job number to be injected into dashboard via AJAX."""
    job_number = (request.GET.get('job_number') or '').strip()
    if not job_number:
        return HttpResponseBadRequest('missing job_number')
    try:
        job = ProductionJob.objects.select_related('product__product_model').get(job_number=job_number)
    except ProductionJob.DoesNotExist:
        return JsonResponse({"ok": False, "html": "<div class='text-red-600'>شماره کار یافت نشد.</div>"})

    logs = (ProductionLog.objects
            .filter(job=job)
            .select_related('user')
            .order_by('logged_at', 'id'))

    # Current/process status
    is_closed = bool(getattr(job, 'finished_at', None))
    current_section = getattr(job, 'current_section', None)
    allowed = [str(s).lower() for s in (getattr(job, 'allowed_sections', []) or [])]
    # Build flow visualization data (ordered subset)
    ORDER = ['assembly','workpage','undercoating','painting','sewing','upholstery','packaging']
    flow = [s for s in ORDER if (not allowed) or (s in allowed)]
    flow_items = [{'slug': s, 'label': dict(SectionChoices.choices).get(s, s)} for s in flow]
    # Build a list of visited sections (unique, order preserved)
    visited_list = []
    for x in logs:
        s = str(x.section).lower()
        if s not in visited_list:
            visited_list.append(s)

    # Flags
    is_scrapped = (getattr(job, 'job_label', '') == 'scrapped') or any(l.is_scrap for l in logs)
    is_deposit = (getattr(job, 'job_label', '') == 'deposit')
    has_external = any(l.is_external for l in logs)

    # Users involved
    users = []
    seen_users = set()
    for l in logs:
        uname = getattr(l.user, 'full_name', None) or getattr(l.user, 'username', None)
        if uname and uname not in seen_users:
            users.append(uname)
            seen_users.add(uname)

    consumption = _compute_job_consumption(job)

    # Jalali date strings for created/finished
    def _to_jalali(dt):
        if not dt:
            return '-'
        try:
            g = timezone.localtime(dt)
            jdt = jdatetime.datetime.fromgregorian(datetime=g)
            return jdt.strftime('%Y-%m-%d %H:%M')
        except Exception:
            return str(dt)
    created_jdt = _to_jalali(getattr(job, 'created_at', None))
    finished_jdt = _to_jalali(getattr(job, 'finished_at', None))

    html = render_to_string('reports/job_details_panel.html', {
        'job': job,
        'logs': logs,
        'is_closed': is_closed,
        'current_section': current_section,
        'flow': flow,
        'flow_items': flow_items,
        'visited': visited_list,
        'is_scrapped': is_scrapped,
        'is_deposit': is_deposit,
        'has_external': has_external,
        'users': users,
        'consumption': consumption,
        'created_jdt': created_jdt,
        'finished_jdt': finished_jdt,
    })
    return JsonResponse({"ok": True, "html": html})


@login_required(login_url="/users/login/")
def job_details_export(request, job_number: str, fmt: str):
    """Export job details to CSV (Excel-friendly) or print-friendly HTML for PDF.

    CSV: returns a text/csv with UTF-8 BOM so Excel opens Persian correctly.
    PDF: returns an HTML with print styles; user can use browser "Save as PDF".
    """
    try:
        job = ProductionJob.objects.select_related('product__product_model').get(job_number=job_number)
    except ProductionJob.DoesNotExist:
        return HttpResponseBadRequest('invalid job number')

    logs = (ProductionLog.objects
            .filter(job=job)
            .select_related('user')
            .order_by('logged_at', 'id'))
    consumption = _compute_job_consumption(job)

    # Jalali converter
    def _to_jalali(dt):
        if not dt:
            return ''
        try:
            g = timezone.localtime(dt)
            jdt = jdatetime.datetime.fromgregorian(datetime=g)
            return jdt.strftime('%Y-%m-%d %H:%M')
        except Exception:
            return str(dt)

    if fmt == 'excel':
        # Show the same styled HTML as PDF (no forced download)
        # English: render the print-friendly template to keep style identical to PDF output.
        created_jdt = _to_jalali(getattr(job, 'created_at', None))
        finished_jdt = _to_jalali(getattr(job, 'finished_at', None))
        html = render_to_string('reports/job_details_export.html', {
            'job': job,
            'logs': logs,
            'consumption': consumption,
            'created_jdt': created_jdt,
            'finished_jdt': finished_jdt,
        })
        return HttpResponse(html)

    if fmt == 'xlsx':
        wb, ws, styles, set_cell, add_title = _prepare_detail_sheet("گزارش جزئیات سفارش", max_cols=6)
        row_idx = 1
        code = (order.subscription_code or str(order.id)).replace('/', '_')
        add_title(row_idx, f"گزارش جزئیات سفارش - {code}")
        row_idx += 2

        def add_kv_row(pairs):
            nonlocal row_idx
            col = 1
            for k, v in pairs:
                set_cell(row_idx, col, k, is_label=True); col += 1
                set_cell(row_idx, col, v); col += 1
            row_idx += 1

        add_kv_row([("نام مشتری", order.customer_name or '-'), ("شهر", order.city or '-')])
        add_kv_row([("مدل", order.model or '-'), ("وضعیت", order.get_status_display())])
        add_kv_row([("کد اشتراک", order.subscription_code or '-'), ("نمایشگاه/فروشگاه", order.exhibition_name or '-')])
        add_kv_row([("تاریخ سفارش", ctx['order_jdt'] or '-'), ("ورود پارچه", ctx['fabric_entry_jdt'] or '-')])
        add_kv_row([("تاریخ تحویل", ctx['delivery_jdt'] or '-')])
        row_idx += 1

        ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=2)
        set_cell(row_idx, 1, "اقلام سفارش", is_label=True, alignment=styles['center_header'])
        row_idx += 1
        item_rows = [[it.get('name') or '', it.get('qty') or ''] for it in items]
        write_table(
            ws,
            headers=["نام آیتم", "تعداد"],
            rows=item_rows,
            start_row=row_idx,
            column_widths=[32, 14],
            table_name="OrderItems",
        )

        return _xlsx_response_from_workbook(wb, f"order_{code}.xlsx")

    # Print-friendly HTML for PDF (render to real PDF if engine is available)
    html = render_to_string('reports/job_details_export.html', {
        'job': job,
        'logs': logs,
        'consumption': consumption,
        'created_jdt': _to_jalali(getattr(job, 'created_at', None)),
        'finished_jdt': _to_jalali(getattr(job, 'finished_at', None)),
    })
    return _pdf_response_from_html(html, f"job_{job.job_number}", request)


@login_required(login_url="/users/login/")
def order_details_panel(request):
    """Return HTML panel for an order to be injected into dashboard via AJAX.

    Accepts one of the following GET parameters:
    - sub: subscription_code (preferred)
    - qr: order QR code string (legacy)
    - id: numeric order id
    If both exist, qr takes precedence.
    """
    badge = (request.GET.get('badge') or '').strip()
    sub = (request.GET.get('sub') or '').strip()
    qr = (request.GET.get('qr') or '').strip()
    oid = (request.GET.get('id') or '').strip()
    base_qs = Order.objects.prefetch_related('items__product')
    order = None

    if oid and oid.isdigit():
        order = base_qs.filter(id=int(oid)).first()

    if not order and badge:
        order = base_qs.filter(badge_number=badge).order_by('-id').first()

    if not order and sub:
        order = base_qs.filter(subscription_code=sub).order_by('-id').first()

    if not order and qr:
        order = base_qs.filter(qr_code=qr).order_by('-id').first()

    if not order and not (badge or sub or qr or oid):
        return HttpResponseBadRequest('missing order key')

    if not order:
        return JsonResponse({"ok": False, "html": "<div class='text-red-600'>سفارش یافت نشد.</div>"})

    # Related jobs linked to this order (via jobs app)
    from jobs.models import ProductionJob
    jobs = list(
        ProductionJob.objects.filter(order=order)
        .select_related('product', 'product__product_model')
        .order_by('-created_at')
    )

    # Collect order items in a simple list for template
    items = []
    for it in getattr(order, 'items', []).all() if hasattr(order, 'items') else []:
        items.append({
            'name': getattr(getattr(it, 'product', None), 'name', '') or '-',
            'qty': getattr(it, 'quantity', 0) or 0,
        })

    # Format Jalali dates as strings similar to job panel style
    def _fmt_jdate(jd):
        try:
            return jd.strftime('%Y-%m-%d') if jd else '-'
        except Exception:
            return str(jd) if jd else '-'

    job_status_map = {}
    job_summary_lines = []
    for job in jobs:
        label = job.get_job_label_display() if hasattr(job, 'get_job_label_display') else getattr(job, 'job_label', '')
        if label:
            job_status_map[label] = job_status_map.get(label, 0) + 1
        product_name = getattr(getattr(job, 'product', None), 'name', '') or ''
        job_summary_lines.append({
            'job_number': getattr(job, 'job_number', '') or '-',
            'status': label or 'نامشخص',
            'product': product_name,
        })
    job_status_badges = list(job_status_map.items())

    status_flow = ['در انتظار', 'در حال ساخت', 'در انبار', 'ارسال شده']
    current_status = getattr(order, 'status', '') or ''
    status_display = ''
    try:
        status_display = order.get_status_display()
    except Exception:
        status_display = current_status

    try:
        status_step_index = status_flow.index(current_status)
    except ValueError:
        status_step_index = None

    status_flow_disabled = status_step_index is None
    status_badge_classes = get_status_badge_classes(status_display)

    inactive_classes = 'bg-white text-gray-400 border-gray-200'
    status_flow_steps = []
    for idx, label in enumerate(status_flow):
        is_current = not status_flow_disabled and idx == status_step_index
        step_class = status_badge_classes + ' border-transparent shadow-sm' if is_current else inactive_classes
        status_flow_steps.append({
            'label': label,
            'class': step_class,
            'is_current': is_current,
        })

    status_summaries = {
        'در انتظار': 'سفارش ثبت شده و در صف برنامه‌ریزی تولید قرار دارد.',
        'در حال ساخت': 'سفارش در حال تولید است و مراحل کارگاهی در جریان است.',
        'در انبار': 'سفارش تولید شده و در انبار برای تحویل یا ارسال آماده است.',
        'ارسال شده': 'سفارش تحویل باربری یا مشتری شده است.',
        'لغو شده': 'سفارش لغو شده و نیاز به پیگیری مجدد ندارد.',
        'گارانتی': 'سفارش در فرآیند خدمات پس از فروش یا گارانتی قرار گرفته است.',
    }
    status_summary_text = status_summaries.get(current_status, f"وضعیت فعلی: {order.get_status_display()}")

    html = render_to_string('reports/order_details_panel.html', {
        'order': order,
        'items': items,
        'jobs': jobs,
        'order_jdt': _fmt_jdate(getattr(order, 'order_date', None)),
        'fabric_entry_jdt': _fmt_jdate(getattr(order, 'fabric_entry_date', None)),
        'delivery_jdt': _fmt_jdate(getattr(order, 'delivery_date', None)),
        'status_flow': status_flow,
        'status_flow_steps': status_flow_steps,
        'status_flow_disabled': status_flow_disabled,
        'status_badge_classes': status_badge_classes,
        'status_summary_text': status_summary_text,
        'job_status_badges': job_status_badges,
        'job_summary_lines': job_summary_lines,
    })
    response_key = {
        'id': str(getattr(order, 'id', '') or ''),
        'badge': str(getattr(order, 'badge_number', '') or ''),
        'sub': str(getattr(order, 'subscription_code', '') or ''),
    }
    return JsonResponse({"ok": True, "html": html, "key": response_key})


@login_required(login_url="/users/login/")
def order_details_export(request, order_id: int, fmt: str):
    """Export order details to Excel-friendly HTML or print-friendly HTML.

    - excel: returns an HTML table that Excel opens cleanly.
    - pdf: returns a print-styled HTML; user can Save as PDF.
    """
    from django.shortcuts import get_object_or_404
    order = get_object_or_404(Order.objects.prefetch_related('items__product'), pk=order_id)

    # Simple lists for template rendering
    items = []
    for it in getattr(order, 'items', []).all() if hasattr(order, 'items') else []:
        items.append({
            'name': getattr(getattr(it, 'product', None), 'name', '') or '-',
            'qty': getattr(it, 'quantity', 0) or 0,
        })

    def _fmt_jdate(jd):
        try:
            return jd.strftime('%Y-%m-%d') if jd else ''
        except Exception:
            return str(jd) if jd else ''

    ctx = {
        'order': order,
        'items': items,
        'order_jdt': _fmt_jdate(getattr(order, 'order_date', None)),
        'fabric_entry_jdt': _fmt_jdate(getattr(order, 'fabric_entry_date', None)),
        'delivery_jdt': _fmt_jdate(getattr(order, 'delivery_date', None)),
    }

    if fmt == 'excel':
        # Show the same styled HTML as PDF (no forced download)
        # English: keep styles identical to PDF by reusing the print template.
        html = render_to_string('reports/order_details_export.html', ctx)
        return HttpResponse(html)

    if fmt == 'xlsx':
        wb, ws, styles, set_cell, add_title = _prepare_detail_sheet("گزارش جزئیات سفارش", max_cols=6)
        row_idx = 1
        code = (order.subscription_code or str(order.id)).replace('/', '_')
        add_title(row_idx, f"گزارش جزئیات سفارش - {code}")
        row_idx += 2

        def add_kv_row(pairs):
            nonlocal row_idx
            col = 1
            for k, v in pairs:
                set_cell(row_idx, col, k, is_label=True); col += 1
                set_cell(row_idx, col, v); col += 1
            row_idx += 1

        add_kv_row([("نام مشتری", order.customer_name or '-'), ("شهر", order.city or '-')])
        add_kv_row([("مدل", order.model or '-'), ("وضعیت", order.get_status_display())])
        add_kv_row([("کد اشتراک", order.subscription_code or '-'), ("نمایشگاه/فروشگاه", order.exhibition_name or '-')])
        add_kv_row([("تاریخ سفارش", ctx['order_jdt'] or '-'), ("ورود پارچه", ctx['fabric_entry_jdt'] or '-')])
        add_kv_row([("تاریخ تحویل", ctx['delivery_jdt'] or '-')])
        row_idx += 1

        ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=2)
        set_cell(row_idx, 1, "اقلام سفارش", is_label=True, alignment=styles['center_header'])
        row_idx += 1
        item_rows = [[it.get('name') or '', it.get('qty') or ''] for it in items]
        write_table(
            ws,
            headers=["نام آیتم", "تعداد"],
            rows=item_rows,
            start_row=row_idx,
            column_widths=[32, 14],
            table_name="OrderItems",
        )

        return _xlsx_response_from_workbook(wb, f"order_{code}.xlsx")

    # Default: print-friendly HTML → render to PDF if possible
    html = render_to_string('reports/order_details_export.html', ctx)
    code = (order.subscription_code or str(order.id)).replace('/', '_')
    return _pdf_response_from_html(html, f"order_{code}", request)


@login_required(login_url="/users/login/")
def log_details_panel(request):
    """Return HTML panel for a single production log entry (for dashboard)."""
    lid = (request.GET.get('id') or '').strip()
    if not lid.isdigit():
        return HttpResponseBadRequest('invalid id')
    try:
        log = ProductionLog.objects.select_related('job', 'user', 'job__product__product_model').get(id=int(lid))
    except ProductionLog.DoesNotExist:
        return JsonResponse({"ok": False, "html": "<div class='text-red-600'>رکورد یافت نشد.</div>"})

    from production_line.models import SectionChoices
    section_label = dict(SectionChoices.choices).get(getattr(log, 'section', ''), getattr(log, 'section', ''))
    job = getattr(log, 'job', None)
    user = getattr(log, 'user', None)

    jdate = getattr(log, 'jdate', None) or ''
    time_str = (getattr(log, 'logged_at', None).strftime('%H:%M') if getattr(log, 'logged_at', None) else '-')

    # Quantities for parts-based logs
    is_parts = bool(getattr(log, 'part_id', None)) and not bool(getattr(log, 'product_id', None))
    produced_qty = int(getattr(log, 'produced_qty', 0) or 0)
    scrap_qty = int(getattr(log, 'scrap_qty', 0) or 0)

    # Build report description text: product/part and flags
    bits = []
    try:
        if getattr(log, 'product', None):
            bits.append(f"محصول: {getattr(log.product, 'name', '')}")
        elif getattr(log, 'part', None):
            bits.append(f"قطعه: {getattr(log.part, 'name', '')}")
        if getattr(log, 'is_scrap', False):
            bits.append('ضایعات')
        if getattr(job, 'job_label', '') == 'scrapped':
            bits.append('اسقاط')
        if getattr(log, 'is_external', False) or getattr(job, 'job_label', '') == 'deposit':
            bits.append('برچسب کلاف بیرون')
    except Exception:
        pass
    report_desc = ' • '.join([b for b in bits if b]) if bits else ''

    # Build badges list for a neat chip row (mimics job panel aesthetics)
    badges = []
    if getattr(log, 'is_scrap', False):
        badges.append('ضایعات')
    if getattr(job, 'job_label', '') == 'scrapped':
        badges.append('اسقاط')
    if getattr(log, 'is_external', False):
        badges.append('خارج از مجموعه')
    if getattr(job, 'job_label', '') == 'deposit':
        badges.append('امانی')

    html = render_to_string('reports/log_details_panel.html', {
        'log': log,
        'job': job,
        'user': user,
        'section_label': section_label,
        'jdate': jdate,
        'time_str': time_str,
        'report_desc': report_desc,
        'badges': badges,
        'is_parts': is_parts,
        'produced_qty': produced_qty,
        'scrap_qty': scrap_qty,
    })
    return JsonResponse({"ok": True, "html": html})


@login_required(login_url="/users/login/")
def log_details_export(request, log_id: int, fmt: str):
    """Export a single production log entry to XLSX or print HTML."""
    from django.shortcuts import get_object_or_404
    log = get_object_or_404(ProductionLog.objects.select_related('job', 'user', 'job__product__product_model'), pk=log_id)
    from production_line.models import SectionChoices
    section_label = dict(SectionChoices.choices).get(getattr(log, 'section', ''), getattr(log, 'section', ''))
    job = getattr(log, 'job', None)
    user = getattr(log, 'user', None)
    jdate = getattr(log, 'jdate', None) or ''
    time_str = (getattr(log, 'logged_at', None).strftime('%H:%M') if getattr(log, 'logged_at', None) else '-')

    is_parts = bool(getattr(log, 'part_id', None)) and not bool(getattr(log, 'product_id', None))
    produced_qty = int(getattr(log, 'produced_qty', 0) or 0)
    scrap_qty = int(getattr(log, 'scrap_qty', 0) or 0)

    ctx = {
        'log': log,
        'job': job,
        'user': user,
        'section_label': section_label,
        'jdate': jdate,
        'time_str': time_str,
        'report_desc': ' • '.join(filter(None, [
            f"محصول: {getattr(log.product, 'name', '')}" if getattr(log, 'product', None) else (f"قطعه: {getattr(log.part, 'name', '')}" if getattr(log, 'part', None) else ''),
            'ضایعات' if getattr(log, 'is_scrap', False) else '',
            'اسقاط' if getattr(job, 'job_label', '') == 'scrapped' else '',
            'برچسب کلاف بیرون' if (getattr(log, 'is_external', False) or getattr(job, 'job_label', '') == 'deposit') else ''
        ])),
        'is_parts': is_parts,
        'produced_qty': produced_qty,
        'scrap_qty': scrap_qty,
    }

    if fmt == 'excel':
        # Show the same styled HTML as PDF (no forced download)
        # English: reuse print-friendly template for identical styling.
        html = render_to_string('reports/log_details_export.html', ctx)
        return HttpResponse(html)

    if fmt == 'xlsx':
        wb, ws, styles, set_cell, add_title = _prepare_detail_sheet("جزئیات ثبت کار", max_cols=6)
        row_idx = 1
        add_title(row_idx, f"جزئیات ثبت کار - {getattr(job, 'job_number', '-')}: {section_label}")
        row_idx += 2

        def add_kv_row(pairs):
            nonlocal row_idx
            col = 1
            for k, v in pairs:
                set_cell(row_idx, col, k, is_label=True); col += 1
                set_cell(row_idx, col, v); col += 1
            row_idx += 1

        add_kv_row([("شماره کار", getattr(job, 'job_number', '-')), ("بخش", section_label)])
        add_kv_row([("تاریخ", jdate), ("ساعت", time_str)])
        add_kv_row([("کاربر", (getattr(user, 'full_name', None) or getattr(user, 'username', '-'))), ("ضایعات", bool(getattr(log, 'is_scrap', False)))])
        add_kv_row([("خارج از مجموعه", bool(getattr(log, 'is_external', False))), ("توضیح", getattr(log, 'note', '') or '-')])
        if is_parts:
            add_kv_row([("تعداد تولید", produced_qty), ("تعداد ضایعات/اسقاط", scrap_qty)])
        row_idx += 1

        return _xlsx_response_from_workbook(wb, f"log_{log.id}.xlsx")

    html = render_to_string('reports/log_details_export.html', ctx)
    return _pdf_response_from_html(html, f"log_{log.id}", request)


@login_required(login_url="/users/login/")
def logs_list_export(request, fmt: str):
    """Export the visible logs list (after client filters) to XLSX or PDF.

    Reads filter parameters from query string to mirror the on-page filters:
    - section: exact SectionChoices key
    - user: user id
    - model: exact model name (case-insensitive)
    - df: from jdate (YYYY/MM/DD or YYYY-MM-DD)
    - dt: to jdate (YYYY/MM/DD or YYYY-MM-DD)
    - q: free text search
    - sort_col, sort_dir: optional sort settings from client (0-based column index)
    """
    from production_line.models import SectionChoices

    def _normalize_digits(value: str | None) -> str:
        """Normalize Persian/Arabic digits to ASCII for consistent comparisons."""
        s = (value or '')
        fa = '۰۱۲۳۴۵۶۷۸۹'
        ar = '٠١٢٣٤٥٦٧٨٩'
        mapping = {fa[i]: str(i) for i in range(10)}
        mapping.update({ar[i]: str(i) for i in range(10)})
        return ''.join(mapping.get(ch, ch) for ch in s)

    def _normalize_search_text(value: str | None) -> str:
        """Lowercase + digit-normalized search text (mirrors frontend normalizeSearchText)."""
        return _normalize_digits((value or '').strip().lower())

    def _to_jdate(s: str | None):
        """Parse Jalali date from 'YYYY/MM/DD' or 'YYYY-MM-DD' with Persian/Arabic digits."""
        s = (s or '').strip()
        if not s:
            return None
        s = _normalize_digits(s).replace('-', '/')
        try:
            parts = [int(x) for x in s.split('/')]
            if len(parts) == 3:
                return jdatetime.date(parts[0], parts[1], parts[2])
        except Exception:
            return None
        return None

    mode = (request.GET.get('mode') or '').strip().lower()

    # ------------------------------
    # Open jobs export (mode=open)
    # ------------------------------
    if mode == 'open':
        ctx = _gather_reports_metrics()
        source = list(ctx.get('logs_open_jobs_list', []))

        sec = (request.GET.get('section') or '').strip()
        mdl = (request.GET.get('model') or '').strip()
        raw_q = (request.GET.get('q') or '').strip()
        q = _normalize_search_text(raw_q)
        df = _to_jdate(request.GET.get('df'))
        dt = _to_jdate(request.GET.get('dt'))

        rows = []
        for item in source:
            if sec and item.get('section') != sec:
                continue
            if mdl:
                model_name = (item.get('model') or '').strip()
                if model_name.lower() != mdl.lower():
                    continue
            created_date = item.get('created_date') or ''
            created_time = item.get('created_time') or ''
            created_dt = None
            if created_date:
                try:
                    created_dt = _to_jdate(created_date)
                except Exception:
                    created_dt = None
            if df and (created_dt is None or created_dt < df):
                continue
            if dt and (created_dt is None or created_dt > dt):
                continue
            if q:
                hay = ' '.join(
                    [
                        str(item.get('job_number') or ''),
                        str(item.get('section_label') or ''),
                        str(item.get('model') or ''),
                        str(item.get('item_name') or ''),
                        str(item.get('label_display') or ''),
                    ]
                )
                if q not in _normalize_search_text(hay):
                    continue
            rows.append([
                item.get('job_number') or '',
                item.get('section_label') or '',
                item.get('model') or '',
                item.get('item_name') or '',
                item.get('label_display') or '',
                f"{created_date} {created_time}".strip(),
            ])

        # Reflect client sort state
        try:
            import re

            def _normalize_digits(s: str) -> str:
                if not s:
                    return ''
                s = str(s)
                s = ''.join(
                    chr(ord('0') + (ord(ch) - ord('۰'))) if '۰' <= ch <= '۹'
                    else chr(ord('0') + (ord(ch) - ord('٠'))) if '٠' <= ch <= '٩'
                    else ch
                    for ch in s
                )
                return s

            scol = int(request.GET.get('sort_col', ''))
            sdir = (request.GET.get('sort_dir') or 'asc').lower()

            def _key_open(row):
                val = row[scol]
                if scol == 0:
                    # Numeric sort on job number (ignore non-digits)
                    txt = _normalize_digits(val)
                    txt = re.sub(r'[^0-9.-]', '', txt)
                    try:
                        return int(txt or 0)
                    except Exception:
                        return 0
                return str(val)

            rows.sort(key=_key_open, reverse=(sdir == 'desc'))
        except Exception:
            pass

        headers = ['شماره کار', 'واحد', 'مدل', 'قطعه/محصول', 'وضعیت', 'تاریخ ایجاد']

        if fmt == 'pdf' or fmt == 'print':
            try:
                gnow = timezone.localtime(timezone.now())
                print_dt = jdatetime.datetime.fromgregorian(datetime=gnow).strftime('%Y/%m/%d %H:%M')
            except Exception:
                print_dt = ''
            html = render_to_string('reports/logs_list_export.html', {
                'title': 'گزارش لیست کارهای باز',
                'headers': headers,
                'rows': rows,
                'print_dt': print_dt,
            })
            if (request.GET.get('dl') or request.GET.get('download')):
                return _pdf_response_from_html(html, 'open_jobs_list', request)
            return HttpResponse(html)

        if fmt == 'xlsx':
            return build_table_response(
                sheet_title="لیست کارهای باز",
                report_title="گزارش لیست کارهای باز",
                headers=headers,
                rows=rows,
                filename="open_jobs_list.xlsx",
                column_widths=[20] * len(headers),
                table_name="OpenJobsTable",
            )

        # Unsupported format for open-mode; fall back to registered mode handling

    # ------------------------------
    # Registered logs export (default)
    # ------------------------------
    qs = (ProductionLog.objects
          .select_related('job', 'user', 'product', 'part')
          .order_by('-logged_at', '-id'))

    sec = (request.GET.get('section') or '').strip()
    uid = (request.GET.get('user') or '').strip()
    mdl = (request.GET.get('model') or '').strip()
    raw_q = (request.GET.get('q') or '').strip()
    q_norm = _normalize_search_text(raw_q)
    df = _to_jdate(request.GET.get('df'))
    dt = _to_jdate(request.GET.get('dt'))

    if sec:
        qs = qs.filter(section=sec)
    if uid.isdigit():
        qs = qs.filter(user_id=int(uid))
    if mdl:
        qs = qs.filter(model__iexact=mdl)
    if df:
        qs = qs.filter(jdate__gte=df)
    if dt:
        qs = qs.filter(jdate__lte=dt)

    # Materialize rows with computed fields exactly like template
    rows = []
    label_map = dict(SectionChoices.choices)
    def _fmt_jdate(d):
        if not d:
            return ''
        try:
            return d.strftime('%Y/%m/%d')
        except Exception:
            try:
                # If it's a datetime/date-like
                import jdatetime as _jd
                if isinstance(d, _jd.date) or hasattr(d, 'year'):
                    return f"{getattr(d,'year', ''):04d}/{getattr(d,'month',''):02d}/{getattr(d,'day',''):02d}"
            except Exception:
                pass
        return str(d)
    for l in qs:
        produced = ''
        scrap = ''
        if getattr(l, 'product_id', None) and not getattr(l, 'part_id', None):
            if getattr(l, 'is_scrap', False):
                scrap = '1'
            else:
                produced = '1'
        else:
            pq = int(getattr(l, 'produced_qty', 0) or 0)
            sq = int(getattr(l, 'scrap_qty', 0) or 0)
            produced = (str(pq) if pq else '')
            scrap = (str(sq) if sq else '')
        jdate = _fmt_jdate(getattr(l, 'jdate', None))
        time_str = getattr(l, 'logged_at', None).strftime('%H:%M') if getattr(l, 'logged_at', None) else ''
        row = [
            (getattr(getattr(l, 'job', None), 'job_number', '') or ''),
            label_map.get(getattr(l, 'section', ''), getattr(l, 'section', '')),
            (getattr(getattr(l, 'user', None), 'full_name', None) or getattr(getattr(l, 'user', None), 'username', '')),
            (getattr(l, 'model', '') or ''),
            (getattr(getattr(l, 'part', None), 'name', None) or getattr(getattr(l, 'product', None), 'name', '') or ''),
            produced,
            scrap,
            f"{jdate} {time_str}".strip(),
        ]
        if q_norm:
            hay = ' '.join(str(v) for v in row)
            if q_norm not in _normalize_search_text(hay):
                continue
        rows.append(row)

    # Optional client-side sort reflection: by column index
    try:
        scol = int(request.GET.get('sort_col', ''))
        sdir = (request.GET.get('sort_dir') or 'asc').lower()
        def _key(row):
            val = row[scol]
            # numeric sort for produced/scrap columns
            if scol in (5, 6):
                try:
                    return int(val or 0)
                except Exception:
                    return 0
            return str(val)
        rows.sort(key=_key, reverse=(sdir == 'desc'))
    except Exception:
        pass

    headers = [
        'شماره کار', 'واحد', 'کاربر', 'مدل', 'قطعه/محصول', 'تعداد تولید شده', 'تعداد ضایعات/اسقاط', 'زمان ثبت'
    ]

    if fmt == 'pdf' or fmt == 'print':
        # Use print-friendly HTML (browser handles Persian perfectly). Auto-print on load.
        # Compose current Jalali timestamp for the print footer/header
        try:
            gnow = timezone.localtime(timezone.now())
            print_dt = jdatetime.datetime.fromgregorian(datetime=gnow).strftime('%Y/%m/%d %H:%M')
        except Exception:
            print_dt = ''
        html = render_to_string('reports/logs_list_export.html', {
            'title': 'گزارش لیست کارهای ثبت‌شده',
            'headers': headers,
            'rows': rows,
            'print_dt': print_dt,
        })
        # If user requested direct download (dl=1), render to PDF server-side
        if (request.GET.get('dl') or request.GET.get('download')):
            return _pdf_response_from_html(html, 'logs_list', request)
        return HttpResponse(html)

    if fmt == 'xlsx':
        return build_table_response(
            sheet_title="لیست کارهای ثبت‌شده",
            report_title="گزارش لیست کارهای ثبت‌شده",
            headers=headers,
            rows=rows,
            filename="logs_list.xlsx",
            column_widths=[20] * len(headers),
            table_name="LogsListTable",
        )

    return HttpResponseBadRequest('invalid format')
