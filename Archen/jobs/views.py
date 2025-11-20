"""Views for the jobs app.

This module provides list, create, edit and bulk delete views for
production jobs.  Access to these views is restricted to managers
and accountants using the same helper present in the ``production_line``
app.  The views mirror the previous implementations in
``production_line.views`` but redirect to the appropriate namespaced
routes in the jobs app.
"""

from __future__ import annotations

from django.contrib.auth.decorators import login_required, user_passes_test
from django.db import transaction
from django.db.models import Q
from django.http import HttpResponseForbidden, Http404
from django.shortcuts import render, redirect
from django.contrib import messages
from django.utils import timezone
import re
from typing import Iterable
from io import BytesIO
import zipfile
import jdatetime

from production_line.views import is_manager_or_accountant
from .models import ProductionJob
from jobs.forms import CreateJobForm
from jobs.services import delete_job_completely, rewind_job_progress
from production_line.models import SectionChoices
from production_line.utils import product_contains_mdf_page
from inventory.models import Product


def _infer_default_allowed_sections(product: Product | None) -> list[str]:
    """Infer default allowed sections based on product BOM (MDF presence).

    - If MDF/page is present in product components: allow all except sewing/upholstery.
    - Otherwise: allow all except workpage.
    This only provides defaults; users can still edit ticks in the form.
    """
    # English comments per project guideline
    if not product:
        return []
    has_mdf = product_contains_mdf_page(product)

    all_sections = [
        SectionChoices.CUTTING,
        SectionChoices.CNC_TOOLS,
        SectionChoices.UNDERCOATING,
        SectionChoices.PAINTING,
        SectionChoices.WORKPAGE,
        SectionChoices.SEWING,
        SectionChoices.UPHOLSTERY,
        SectionChoices.ASSEMBLY,
        SectionChoices.PACKAGING,
    ]
    if has_mdf:
        # Exclude sewing and upholstery when MDF/page is present
        return [s for s in map(str, all_sections) if s not in (SectionChoices.SEWING, SectionChoices.UPHOLSTERY)]
    # Else exclude workpage
    return [s for s in map(str, all_sections) if s != SectionChoices.WORKPAGE]


PRODUCT_SECTION_FLOW = [
    str(SectionChoices.ASSEMBLY),
    str(SectionChoices.WORKPAGE),
    str(SectionChoices.UNDERCOATING),
    str(SectionChoices.PAINTING),
    str(SectionChoices.SEWING),
    str(SectionChoices.UPHOLSTERY),
    str(SectionChoices.PACKAGING),
]

SECTION_LABEL_MAP = {str(slug): label for slug, label in SectionChoices.choices}


def _ordered_allowed_sections(raw_sections: Iterable[str] | None) -> list[str]:
    """Return allowed sections normalized to the business flow order."""
    if not raw_sections:
        return []
    raw_set = {str(sec).lower() for sec in raw_sections if sec}
    return [slug for slug in PRODUCT_SECTION_FLOW if slug in raw_set]


def _build_progress_state(job: ProductionJob | None, raw_sections: Iterable[str] | None) -> dict:
    """Build metadata for the allowed-section highlight component."""
    flow = _ordered_allowed_sections(raw_sections)
    cursor = 0
    logged_sections: set[str] = set()

    if job and flow:
        try:
            qs = job.productionlog_set.filter(section__in=flow)
            logged_sections = {
                str(sec).lower()
                for sec in qs.values_list('section', flat=True)
            }
        except Exception:
            logged_sections = set()
        for slug in flow:
            if slug in logged_sections:
                cursor += 1
            else:
                break

    highlight_slug = flow[cursor] if cursor < len(flow) else None

    items = []
    for idx, slug in enumerate(flow):
        if idx < cursor:
            state = 'done'
        elif idx == cursor:
            state = 'active'
        else:
            state = 'pending'
        items.append({
            'slug': slug,
            'label': SECTION_LABEL_MAP.get(slug, slug),
            'index': idx,
            'state': state,
            'can_jump': bool(job) and idx < cursor,
        })

    return {
        'items': items,
        'cursor': cursor,
        'highlight_slug': highlight_slug,
        'flow': flow,
        'flow_length': len(flow),
    }


@login_required
@user_passes_test(is_manager_or_accountant)
def job_list_view(request):
    """Display the list of jobs with optional filtering and searching."""
    label_filter = (request.GET.get('label') or '').strip()
    search_query = (request.GET.get('search') or '').strip()
    qs_all = ProductionJob.objects.all().select_related('product', 'part')
    qs = qs_all
    if label_filter:
        qs = qs.filter(job_label=label_filter)
    if search_query:
        qs = qs.filter(
            Q(job_number__icontains=search_query) |
            Q(product__name__icontains=search_query) |
            Q(part__name__icontains=search_query)
        )
    jobs = qs.order_by('-created_at')
    active_jobs_count = qs.filter(finished_at__isnull=True).count()
    return render(request, 'jobs/job_list.html', {
        'jobs': jobs,
        'label_choices': ProductionJob.LABEL_CHOICES,
        'current_label': label_filter,
        'search_query': search_query,
        'active_jobs_count': active_jobs_count,
        # Total count of jobs (unfiltered) for status bar display
        'jobs_total': qs_all.count(),
        # Explicit alias for clarity in templates/JS when distinguishing filtered vs. all
        'jobs_total_all': qs_all.count(),
    })


@login_required
@user_passes_test(is_manager_or_accountant)
def jobs_list_export_xlsx(request):
    """Export filtered jobs list to XLSX, using a real XLSX library to avoid corrupt content."""
    label_filter = (request.GET.get('label') or '').strip()
    search_query = (request.GET.get('search') or '').strip()

    qs = ProductionJob.objects.all().select_related('product', 'part', 'order')
    if label_filter:
        qs = qs.filter(job_label=label_filter)
    if search_query:
        qs = qs.filter(
            Q(job_number__icontains=search_query) |
            Q(product__name__icontains=search_query) |
            Q(part__name__icontains=search_query)
        )

    jobs = list(qs.order_by('-created_at'))

    def fmt_dt(value):
        if not value:
            return ''
        try:
            gdt = timezone.localtime(value)
            jdt = jdatetime.datetime.fromgregorian(datetime=gdt)
            return jdt.strftime('%Y/%m/%d %H:%M')
        except Exception:
            return str(value)

    def job_label_display(job: ProductionJob) -> str:
        try:
            return job.get_job_label_display()
        except Exception:
            return job.job_label or ''

    def job_stage_display(job: ProductionJob) -> str:
        display = ''
        try:
            display = job.get_current_section_display()
        except Exception:
            display = ''
        if display:
            return display
        raw = getattr(job, 'current_section', None)
        if raw:
            try:
                return SectionChoices(raw).label
            except Exception:
                return str(raw)
        return 'ثبت نشده'

    def job_account_display(job: ProductionJob) -> str:
        value = getattr(job, 'deposit_account', '') or ''
        return value.strip()

    headers = [
        'شماره کار',
        'برچسب کار',
        'مرحله فعلی خط تولید',
        'طرف حساب',
        'مدل',
        'محصول',
        'تاریخ ایجاد',
        'تاریخ بسته شدن',
    ]

    rows: list[list[str]] = []
    for job in jobs:
        if job.product:
            model_name = getattr(getattr(job.product, 'product_model', None), 'name', '') or ''
            product_name = job.product.name or ''
        else:
            model_name = ''
            product_name = ''
        rows.append([
            job.job_number or '',
            job_label_display(job),
            job_stage_display(job),
            job_account_display(job),
            model_name,
            product_name,
            fmt_dt(job.created_at),
            fmt_dt(job.finished_at),
        ])

    # Build XLSX via openpyxl to avoid low-level XML issues
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        from openpyxl.utils import get_column_letter
        from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
    except ImportError:
        from django.http import HttpResponseServerError
        return HttpResponseServerError("کتابخانه openpyxl نصب نشده است؛ لطفاً با مدیر سیستم تماس بگیرید.")

    wb = Workbook()
    ws = wb.active
    ws.title = "لیست کارها"
    try:
        ws.sheet_view.rightToLeft = True
    except Exception:
        pass

    title_font = Font(name="Tahoma", bold=True, size=14)
    header_font = Font(name="Tahoma", bold=True, size=11)
    cell_font = Font(name="Tahoma", size=11)
    center_header = Alignment(horizontal="center", vertical="center")
    right_cell = Alignment(horizontal="right", vertical="center", wrap_text=True)
    header_fill = PatternFill("solid", fgColor="FFF9FAFB")

    row_idx = 1

    # Title
    title_text = "گزارش لیست کارها"
    ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=len(headers))
    c = ws.cell(row=row_idx, column=1, value=title_text)
    c.font = title_font
    c.alignment = center_header
    row_idx += 1

    # Subtitle with Jalali timestamp (like orders list)
    try:
        gnow = timezone.localtime(timezone.now())
        generated = jdatetime.datetime.fromgregorian(datetime=gnow).strftime('%Y/%m/%d %H:%M')
    except Exception:
        generated = ''
    subtitle = f"تاریخ تهیه: {generated}" if generated else ''
    if subtitle:
        ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=len(headers))
        c = ws.cell(row=row_idx, column=1, value=subtitle)
        c.font = cell_font
        c.alignment = right_cell
        row_idx += 1

    # Header row
    for col_idx, label in enumerate(headers, start=1):
        c = ws.cell(row=row_idx, column=col_idx, value=label)
        c.font = header_font
        c.alignment = center_header
        c.fill = header_fill
    header_row_idx = row_idx
    row_idx += 1

    # Data rows
    for data_row in rows:
        for col_idx, raw_value in enumerate(data_row, start=1):
            text = '' if raw_value is None else str(raw_value)
            text = ILLEGAL_CHARACTERS_RE.sub('', text)
            c = ws.cell(row=row_idx, column=col_idx, value=text)
            c.font = cell_font
            # Job number and dates: keep as LTR inside RTL sheet using right alignment
            c.alignment = right_cell
        row_idx += 1

    # Basic column widths similar to manual export
    for col_idx in range(1, len(headers) + 1):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = 24

    # Thin borders for entire table (optional: keeps visual parity with orders export)
    try:
        from openpyxl.styles import Border, Side

        thin_side = Side(style="thin", color="FFE5E7EB")
        border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
        for row in ws.iter_rows(min_row=header_row_idx, max_row=row_idx - 1, min_col=1, max_col=len(headers)):
            for cell in row:
                if cell.border is None or cell.border == Border():
                    cell.border = border
    except Exception:
        pass

    bio = BytesIO()
    wb.save(bio)

    from django.http import HttpResponse

    resp = HttpResponse(
        bio.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    resp['Content-Disposition'] = "attachment; filename=jobs_list.xlsx"
    return resp


@login_required
@user_passes_test(is_manager_or_accountant)
def job_add_view(request):
    """Create a new job using the generic CreateJobForm."""
    if request.method == 'POST':
        form = CreateJobForm(request.POST)
        if form.is_valid():
            job_number_val = form.cleaned_data['job_number']
            product = form.cleaned_data.get('product')
            allowed_sections = form.cleaned_data.get('allowed_sections') or []
            allowed_sections_slugs = [str(sec) for sec in allowed_sections]
            job_label = form.cleaned_data.get('job_label') or 'in_progress'
            deposit_account = (form.cleaned_data.get('deposit_account') or '').strip()

            # Prevent duplicate job numbers while the original row exists.
            # English: If a job with this number already exists (not deleted),
            # block creation and show a clear Persian warning.
            if ProductionJob.objects.filter(job_number=job_number_val).exists():
                form.add_error('job_number', "این شماره کار قبلاً ثبت شده است و تا زمان حذف، امکان ثبت مجدد ندارد.")
                progress_payload = _build_progress_state(None, form['allowed_sections'].value())
                return render(request, 'jobs/job_form.html', {
                    'form': form,
                    'is_editing': False,
                    'section_progress_items': progress_payload['items'],
                    'section_progress_cursor': progress_payload['cursor'],
                    'section_progress_length': progress_payload['flow_length'],
                    'section_progress_highlight': progress_payload['highlight_slug'],
                })

            # Determine the initial status based on the label
            # English: 'repaired' is a label for repair-type jobs, but they are still in progress
            # until the last allowed section is logged. So start them as in_progress.
            if job_label in ('in_progress', 'deposit', 'repaired'):
                initial_status = 'in_progress'
            elif job_label == 'warranty':
                initial_status = 'warranty'
            else:
                # completed/scrapped set explicit closure state
                initial_status = job_label
            # Build object with all fields before first save to avoid update_fields on new instance
            job_obj = ProductionJob(
                job_number=job_number_val,
                product=(product if product else None),
                status=initial_status,
                job_label=job_label,
                deposit_account=(deposit_account if job_label == 'deposit' else None),
            )
            # Persist allowed sections on the job (do not clear if user chose explicitly)
            if hasattr(job_obj, 'allowed_sections'):
                job_obj.allowed_sections = allowed_sections_slugs
            # Ensure only one job is flagged as default at a time, then mark this one
            ProductionJob.objects.filter(is_default=True).exclude(job_number=job_number_val).update(is_default=False)
            if hasattr(job_obj, 'is_default'):
                job_obj.is_default = True
            # If created as scrapped/completed, set finished_at immediately
            if job_obj.status == 'scrapped' and not job_obj.finished_at:
                job_obj.finished_at = timezone.now()
            if job_obj.status == 'completed' and not job_obj.finished_at:
                job_obj.finished_at = timezone.now()
            # Save once (insert)
            job_obj.save()
            messages.success(request, "کار جدید با موفقیت ایجاد شد.")
            return redirect('jobs:job_list')
    else:
        # Provide intelligent defaults when product is preselected via querystring
        initial = {}
        prod_id = request.GET.get('product')
        if prod_id:
            try:
                product = Product.objects.get(pk=prod_id)
                initial['product'] = product.pk
                initial['allowed_sections'] = _infer_default_allowed_sections(product)
            except Product.DoesNotExist:
                pass
        form = CreateJobForm(initial=initial)
    progress_payload = _build_progress_state(None, form['allowed_sections'].value())
    return render(request, 'jobs/job_form.html', {
        'form': form,
        'is_editing': False,
        'section_progress_items': progress_payload['items'],
        'section_progress_cursor': progress_payload['cursor'],
        'section_progress_length': progress_payload['flow_length'],
        'section_progress_highlight': progress_payload['highlight_slug'],
    })


@login_required
@user_passes_test(is_manager_or_accountant)
def job_edit_view(request, pk: int):
    """Edit an existing job.

    If the job cannot be found a 404 is raised.  On successful
    submission the user is redirected back to the jobs list.
    """
    try:
        job = ProductionJob.objects.select_related('product', 'part').get(pk=pk)
    except ProductionJob.DoesNotExist:
        raise Http404("کار مورد نظر یافت نشد.")
    stored_sections = getattr(job, 'allowed_sections', []) or []
    progress_payload = _build_progress_state(job, stored_sections)
    current_flow = progress_payload['flow']
    current_cursor = progress_payload['cursor']
    flow_length = progress_payload['flow_length']

    if request.method == 'POST':
        form = CreateJobForm(request.POST)
        if form.is_valid():
            product = form.cleaned_data.get('product')
            allowed_sections = form.cleaned_data.get('allowed_sections') or []
            allowed_sections_slugs = [str(sec) for sec in allowed_sections]
            job_label = form.cleaned_data.get('job_label') or 'in_progress'
            deposit_account = (form.cleaned_data.get('deposit_account') or '').strip()

            requested_cursor_raw = request.POST.get('progress_cursor', current_cursor)
            try:
                requested_cursor = int(requested_cursor_raw)
            except (TypeError, ValueError):
                requested_cursor = current_cursor
            requested_cursor = max(0, min(requested_cursor, current_cursor))

            updated_fields: set[str] = set()
            removed_logs = 0
            new_current_section = job.current_section

            if job.product != (product if product else None):
                job.product = product if product else None
                updated_fields.add('product')

            job.job_label = job_label
            updated_fields.add('job_label')
            if job_label == 'deposit':
                job.status = 'in_progress'
                job.deposit_account = deposit_account
                updated_fields.update({'status', 'deposit_account'})
            elif job_label in ('in_progress', 'repaired'):
                # English: Keep repair-type jobs in progress until completion logic closes them
                job.status = 'in_progress'
                job.deposit_account = None
                updated_fields.update({'status', 'deposit_account'})
            elif job_label == 'warranty':
                job.status = 'warranty'
                job.deposit_account = None
                updated_fields.update({'status', 'deposit_account'})
            else:
                # completed/scrapped
                job.status = job_label
                job.deposit_account = None
                updated_fields.update({'status', 'deposit_account'})
            # Do not force-clear allowed sections for non in-progress labels.
            # Defaults برای این برچسب‌ها در فرم خالی است اما انتخاب کاربر باید حفظ شود.

            # Persist allowed sections
            if hasattr(job, 'allowed_sections'):
                job.allowed_sections = allowed_sections_slugs
                updated_fields.add('allowed_sections')

            with transaction.atomic():
                if current_flow and requested_cursor < current_cursor:
                    removed_logs, new_current_section = rewind_job_progress(
                        job,
                        current_flow,
                        requested_cursor,
                        current_cursor,
                    )
                    current_cursor = requested_cursor

                desired_current = new_current_section if current_flow else job.current_section
                if job.current_section != desired_current:
                    job.current_section = desired_current
                    updated_fields.add('current_section')

                if current_flow and current_cursor < flow_length:
                    if job.finished_at:
                        job.finished_at = None
                        updated_fields.add('finished_at')
                    if job.status in ('scrapped', 'completed'):
                        job.status = 'in_progress'
                        updated_fields.add('status')
                else:
                    if job.status in ('scrapped', 'completed'):
                        if not job.finished_at:
                            job.finished_at = timezone.now()
                            updated_fields.add('finished_at')
                    else:
                        if job.finished_at:
                            job.finished_at = None
                            updated_fields.add('finished_at')

                if updated_fields:
                    job.save(update_fields=list(updated_fields))

            if removed_logs and current_flow:
                if current_cursor < flow_length:
                    next_label = SECTION_LABEL_MAP.get(current_flow[current_cursor], current_flow[current_cursor])
                    messages.info(
                        request,
                        f"{removed_logs} ثبت روزانه حذف شد و کار دوباره در فهرست بخش «{next_label}» قرار گرفت."
                    )
                else:
                    last_label = SECTION_LABEL_MAP.get(current_flow[-1], current_flow[-1])
                    messages.info(
                        request,
                        f"{removed_logs} ثبت روزانه حذف شد و وضعیت کار برای بخش «{last_label}» بدون تغییر ماند."
                    )
            messages.success(request, "کار با موفقیت ویرایش شد.")
            return redirect('jobs:job_list')
    else:
        initial = {
            'job_number': job.job_number,
            'model': (job.product.product_model.name if job.product and job.product.product_model else ''),
            'product': (job.product.pk if job.product else None),
            'allowed_sections': getattr(job, 'allowed_sections', []) or [],
            'job_label': job.job_label,
            'deposit_account': job.deposit_account or '',
        }
        form = CreateJobForm(initial=initial)
    return render(request, 'jobs/job_form.html', {
        'form': form,
        'is_editing': True,
        'job': job,
        'section_progress_items': progress_payload['items'],
        'section_progress_cursor': progress_payload['cursor'],
        'section_progress_length': progress_payload['flow_length'],
        'section_progress_highlight': progress_payload['highlight_slug'],
    })


@login_required
@user_passes_test(is_manager_or_accountant)
def job_bulk_delete_view(request):
    """Delete multiple jobs at once via POST."""
    if request.method != 'POST':
        return HttpResponseForbidden("درخواست نامعتبر است.")
    ids = request.POST.getlist('ids')
    qs = ProductionJob.objects.filter(pk__in=ids)
    if not qs.exists():
        messages.warning(request, "هیچ کار انتخاب نشده بود.")
        return redirect('jobs:job_list')

    deleted_jobs = 0
    removed_logs = 0
    failed_jobs: list[str] = []

    jobs_to_delete = list(qs.select_related('product', 'part'))

    for job in jobs_to_delete:
        try:
            logs_deleted, job_deleted = delete_job_completely(job)
            deleted_jobs += job_deleted
            removed_logs += logs_deleted
        except Exception:
            failed_jobs.append(job.job_number)

    if deleted_jobs:
        if removed_logs:
            messages.success(
                request,
                f"{deleted_jobs} کار به همراه {removed_logs} گزارش روزانه حذف و موجودی‌ها به حالت قبل برگشتند."
            )
        else:
            messages.success(request, f"{deleted_jobs} کار بدون گزارش حذف شد.")

    if failed_jobs:
        preview = ', '.join(failed_jobs[:5])
        extra = " و ..." if len(failed_jobs) > 5 else ""
        messages.error(request, f"حذف این شماره کارها انجام نشد: {preview}{extra}.")
    return redirect('jobs:job_list')
