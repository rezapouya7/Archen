# PATH: /Archen/orders/views.py
# -*- coding: utf-8 -*-
# Archen/orders/views.py

import json
from io import BytesIO
import jdatetime
from django.views.generic import ListView, CreateView, UpdateView  # type: ignore # noqa: E501
from django.urls import reverse_lazy, reverse  # type: ignore
from django.http import JsonResponse  # type: ignore
from django.http import HttpResponse
from django.views import View
from django.shortcuts import redirect, render, get_object_or_404
from django.contrib import messages
from django.db import IntegrityError, transaction
from django.db.models import F
from django.utils import timezone
from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin

from .models import Order, OrderItem
from .forms import OrderForm, _extract_product_ids
from inventory.models import Product
from inventory.models import Part  # update part inventory on order create/update/delete # noqa: E501

from production_line.models import ProductStock  # <-- added
from production_line.models import ProductionLog
from django.db import models  # for Q in job selection logic

from jobs.models import ProductionJob
from jobs.views import _build_progress_state


def _generate_unique_qr_code(*, exclude_pk=None, max_attempts=6):
    """Return a QR code not used by any existing order."""
    import uuid
    for _ in range(max_attempts):
        candidate = uuid.uuid4().hex
        qs = Order.objects.filter(qr_code=candidate)
        if exclude_pk:
            qs = qs.exclude(pk=exclude_pk)
        if not qs.exists():
            return candidate
    raise RuntimeError("Unable to generate a unique QR code after several attempts.")


def _ensure_unique_qr_code(preferred=None, *, exclude_pk=None):
    """
    Use ``preferred`` if it is free; otherwise fall back to a freshly generated
    unique QR code.
    """
    candidate = (preferred or '').strip()
    if candidate:
        qs = Order.objects.filter(qr_code=candidate)
        if exclude_pk:
            qs = qs.exclude(pk=exclude_pk)
        if not qs.exists():
            return candidate
    return _generate_unique_qr_code(exclude_pk=exclude_pk)

class OrderListView(LoginRequiredMixin, ListView):
    login_url = "/users/login/"
    model = Order
    template_name = 'orders/orders_list.html'
    context_object_name = 'orders'
    ordering = ['-id']
    # Disable pagination to show all orders on the list page
    # This aligns with the client-side search/sort and the request to show all rows
    paginate_by = None

    def get_queryset(self):
        """
        Filter orders by optional ``status`` and apply server-side search when
        a ``search`` query parameter is provided. This enables dynamic, live
        searching across stored fields (database-backed) instead of only
        filtering the visible DOM on the client.
        """
        # Base queryset with related joins for efficient access in templates.
        qs = super().get_queryset().prefetch_related('items__product__product_model')
        status_filter = (self.request.GET.get('status') or '').strip()
        if status_filter:
            qs = qs.filter(status=status_filter)
        # Apply server-side search across multiple fields if provided
        search_raw = (self.request.GET.get('search') or '').strip()
        if search_raw:
            from django.db.models import Q
            # Normalize whitespace
            qstr = ' '.join(search_raw.split())

            # Helper to create digit-variant strings for broader matching.
            def _variants(s: str) -> list[str]:
                # Returns [raw, ascii_digits, persian_digits, arabic_indic_digits]
                try:
                    raw = s
                    # Map Persian/Arabic-Indic → ASCII
                    def to_ascii_digits(t: str) -> str:
                        out = []
                        for ch in t:
                            code = ord(ch)
                            if 0x06F0 <= code <= 0x06F9:
                                out.append(chr(code - 0x06F0 + ord('0')))
                            elif 0x0660 <= code <= 0x0669:
                                out.append(chr(code - 0x0660 + ord('0')))
                            else:
                                out.append(ch)
                        return ''.join(out)

                    ascii_s = to_ascii_digits(raw)
                    # Map ASCII → Persian and ASCII → Arabic-Indic
                    def map_digits(t: str, base: int) -> str:
                        out = []
                        for ch in t:
                            if '0' <= ch <= '9':
                                out.append(chr(ord(ch) - 48 + base))
                            else:
                                out.append(ch)
                        return ''.join(out)

                    persian = map_digits(ascii_s, 0x06F0)
                    arabic = map_digits(ascii_s, 0x0660)
                    return [raw, ascii_s, persian, arabic]
                except Exception:
                    return [s]

            variants = _variants(qstr)

            # Build a broad OR query across key text fields and related data.
            # Limit search to fields exposed on the order form so results never
            # reference hidden/internal data (e.g., job numbers).
            fields = [
                'badge_number__icontains',
                'subscription_code__icontains',
                'customer_name__icontains',
                'exhibition_name__icontains',
                'city__icontains',
                'producer__icontains',
                'customer_phone__icontains',
                'driver_phone__icontains',
                'sender__icontains',
                'driver_name__icontains',
                'fabric_description__icontains',
                'fabric_code__icontains',
                'color_code__icontains',
                'description__icontains',
                'status__icontains',
                'current_stage__icontains',
            ]

            search_q = Q()
            for v in variants:
                for f in fields:
                    search_q |= Q(**{f: v})

            # Attempt a date match for Jalali inputs like 1402/07/20 → Gregorian
            # If parse succeeds, filter equality on date fields.
            try:
                import re
                import jdatetime
                m = re.match(r"^(\d{4})[\-/](\d{1,2})[\-/](\d{1,2})$", variants[1])  # ASCII-digits variant
                if m:
                    jy, jm, jd = (int(m.group(1)), int(m.group(2)), int(m.group(3)))
                    gdate = jdatetime.date(jy, jm, jd).togregorian()
                    search_q |= (
                        Q(order_date=gdate) |
                        Q(delivery_date=gdate) |
                        Q(fabric_entry_date=gdate)
                    )
            except Exception:
                # Ignore date parsing errors; text fields cover most searches.
                pass

            qs = qs.filter(search_q).distinct()
        return qs

    def get_context_data(self, **kwargs):
        """
        Supply additional context for the orders list template, including the
        available status choices and the currently selected status.  The
        ``orders`` key is provided by ListView and will reflect any filtering.
        """
        context = super().get_context_data(**kwargs)
        # Provide status choices for the filter dropdown
        context['status_choices'] = Order.STATUS_CHOICES
        context['stage_choices'] = [choice[0] for choice in Order.STAGE_CHOICES]
        # Current selected status for marking as selected in the template
        context['current_status'] = (self.request.GET.get('status') or '').strip()
        # Pass through search query so the value persists in the input box
        context['search_query'] = (self.request.GET.get('search') or '').strip()
        # Total count of orders (unfiltered) for status bar display
        try:
            context['orders_total'] = Order.objects.count()
        except Exception:
            context['orders_total'] = 0
        return context

    def render_to_response(self, context, **response_kwargs):
        """Render the full list page always.

        We switched client-side logic to fetch the full page and extract the
        updated tbody + status bar to avoid partial-template issues. Keeping
        server response uniform eliminates template parsing mismatches.
        """
        return super().render_to_response(context, **response_kwargs)


def orders_list_export_xlsx(request):
    """
    Export filtered orders to XLSX. The export mirrors the list filters and
    includes every field surfaced on the edit form in addition to the list
    columns so downstream users get the complete record.
    """
    list_view = OrderListView()
    list_view.request = request
    list_view.args = ()
    list_view.kwargs = {}
    orders = list(list_view.get_queryset())
    order_ids = [o.id for o in orders if getattr(o, 'id', None)]

    jobs_map: dict[int, list[str]] = {}
    if order_ids:
        try:
            from jobs.models import ProductionJob
            job_rows = (ProductionJob.objects
                        .filter(order_id__in=order_ids)
                        .values_list('order_id', 'job_number')
                        .order_by('order_id', 'job_number'))
            for oid, job_number in job_rows:
                jobs_map.setdefault(int(oid), []).append(job_number or '')
        except Exception:
            jobs_map = {}

    def fmt_date(value):
        if not value:
            return ''
        try:
            return value.strftime('%Y/%m/%d')
        except Exception:
            try:
                return jdatetime.date(value.year, value.month, value.day).strftime('%Y/%m/%d')
            except Exception:
                return str(value)

    def join_clean(items):
        cleaned = [str(it).strip() for it in items if str(it).strip()]
        return '، '.join(cleaned)

    rows: list[list[str]] = []
    for order in orders:
        store_name = getattr(order, 'exhibition_name', '') or getattr(order, 'store_name', '') or ''
        status_label = order.get_status_display() if hasattr(order, 'get_status_display') else getattr(order, 'status', '')
        stage_label = order.get_current_stage_display() if hasattr(order, 'get_current_stage_display') else getattr(order, 'current_stage', '')
        model_tokens = [tok.strip() for tok in (getattr(order, 'model', '') or '').split(',') if tok.strip()]
        items_summary: list[str] = []
        try:
            for item in order.items.all():
                product_name = getattr(getattr(item, 'product', None), 'name', '') or ''
                if not product_name and getattr(getattr(item, 'product', None), 'product_model', None):
                    try:
                        product_name = item.product.product_model.name or ''
                    except Exception:
                        product_name = ''
                product_name = product_name or f"محصول #{getattr(item, 'product_id', '')}"
                qty = getattr(item, 'quantity', '') or ''
                label = f"{product_name} x{qty}" if qty else product_name
                items_summary.append(label.strip())
        except Exception:
            items_summary = []
        job_numbers = jobs_map.get(getattr(order, 'id', None), [])

        rows.append([
            getattr(order, 'badge_number', '') or '',
            getattr(order, 'customer_name', '') or '',
            store_name,
            fmt_date(getattr(order, 'order_date', None)),
            status_label,
            stage_label,
            getattr(order, 'subscription_code', '') or '',
            getattr(order, 'producer', '') or '',
            getattr(order, 'customer_phone', '') or '',
            getattr(order, 'city', '') or '',
            getattr(order, 'driver_phone', '') or '',
            getattr(order, 'sender', '') or '',
            getattr(order, 'driver_name', '') or '',
            fmt_date(getattr(order, 'delivery_date', None)),
            fmt_date(getattr(order, 'fabric_entry_date', None)),
            getattr(order, 'fabric_code', '') or '',
            getattr(order, 'color_code', '') or '',
            getattr(order, 'fabric_description', '') or '',
            getattr(order, 'description', '') or '',
            join_clean(model_tokens),
            join_clean(items_summary),
            join_clean(job_numbers),
        ])

    headers = [
        'شماره بیجک',
        'نام مشتری',
        'نام فروشگاه',
        'تاریخ سفارش',
        'وضعیت',
        'مرحله فعلی',
        'کد اشتراک مشتری',
        'تولید کننده',
        'شماره تماس مشتری',
        'شهر',
        'شماره تماس راننده',
        'ارسال کننده',
        'نام راننده',
        'تاریخ تحویل',
        'تاریخ ورود پارچه',
        'کد پارچه',
        'کد رنگ',
        'توضیحات پارچه',
        'توضیحات سفارش',
        'مدل‌های انتخاب‌شده',
        'محصولات انتخاب‌شده',
        'شماره کارها',
    ]

    # Build XLSX using openpyxl (same approach as jobs list export)
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
    except ImportError:
        return HttpResponse("کتابخانه openpyxl نصب نشده است؛ لطفاً با مدیر سیستم تماس بگیرید.", status=500)

    wb = Workbook()
    ws = wb.active
    ws.title = "لیست سفارش‌ها"
    try:
        ws.sheet_view.rightToLeft = True
    except Exception:
        pass

    title_font = Font(name="Tahoma", bold=True, size=14)
    header_font = Font(name="Tahoma", bold=True, size=11)
    cell_font = Font(name="Tahoma", size=11)
    center_header = Alignment(horizontal="center", vertical="center")
    right_cell = Alignment(horizontal="right", vertical="center", wrap_text=True)
    header_fill = PatternFill("solid", fgColor="FFF9FAFB")

    row_idx = 1

    # Title
    title_text = "گزارش لیست سفارش‌ها"
    ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=len(headers))
    c = ws.cell(row=row_idx, column=1, value=title_text)
    c.font = title_font
    c.alignment = center_header
    row_idx += 1

    # Subtitle with Jalali timestamp (same logic as before)
    try:
        gnow = timezone.localtime(timezone.now())
        print_dt = jdatetime.datetime.fromgregorian(datetime=gnow).strftime('%Y/%m/%d %H:%M')
    except Exception:
        print_dt = ''
    subtitle = f"تاریخ تهیه: {print_dt}" if print_dt else ''
    if subtitle:
        ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=len(headers))
        c = ws.cell(row=row_idx, column=1, value=subtitle)
        c.font = cell_font
        c.alignment = right_cell
        row_idx += 1

    # Header row
    for col_idx, label in enumerate(headers, start=1):
        c = ws.cell(row=row_idx, column=col_idx, value=label)
        c.font = header_font
        c.alignment = center_header
        c.fill = header_fill
    header_row_idx = row_idx
    row_idx += 1

    # Data rows
    for data_row in rows:
        for col_idx, raw_value in enumerate(data_row, start=1):
            text = '' if raw_value is None else str(raw_value)
            text = ILLEGAL_CHARACTERS_RE.sub('', text)
            c = ws.cell(row=row_idx, column=col_idx, value=text)
            c.font = cell_font
            c.alignment = right_cell
        row_idx += 1

    # Column widths similar to previous export
    for col_idx in range(1, len(headers) + 1):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = 24

    # Thin borders for the whole data table
    thin_side = Side(style="thin", color="FFE5E7EB")
    border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    for row in ws.iter_rows(min_row=header_row_idx, max_row=row_idx - 1, min_col=1, max_col=len(headers)):
        for cell in row:
            if cell.border is None or cell.border == Border():
                cell.border = border

    bio = BytesIO()
    wb.save(bio)

    resp = HttpResponse(
        bio.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    resp['Content-Disposition'] = "attachment; filename=orders_list.xlsx"
    return resp


@login_required(login_url="/users/login/")
def warranty_card(request):
    """Render the print-ready warranty card template.

    If an order QR image URL is available via querystring (e.g., ?qr=...), pass
    it to the template as ``order.qr_image_url``-compatible context.
    """
    qr = (request.GET.get('qr') or '').strip()
    order_id = request.GET.get('order')
    ctx = {}
    if order_id:
        order = get_object_or_404(Order, pk=order_id)
        ctx['order'] = order
    elif qr:
        ctx['order'] = type('OrderCtx', (), {'qr_image_url': qr})()
    # Accept direct data URL for QR image from the form page to ensure exact match
    qr_img = (request.GET.get('qr_img') or '').strip()
    if qr_img:
        # Force use this image URL in template
        if 'order' in ctx and ctx['order'] is not None:
            setattr(ctx['order'], 'qr_image_url', qr_img)
        else:
            ctx['order'] = type('OrderCtx', (), {'qr_image_url': qr_img})()
    # Optional field overrides from query params to reflect unsaved edits in the form.
    # This allows the warranty card to mirror current inputs without requiring a DB save.
    try:
        ov_exhibition = (request.GET.get('exhibition_name') or '').strip()
        ov_customer = (request.GET.get('customer_name') or '').strip()
        ov_badge = (request.GET.get('badge_number') or '').strip()
        ov_order_date = (request.GET.get('order_date') or '').strip()
        ov_delivery_date = (request.GET.get('delivery_date') or '').strip()
        if 'order' in ctx and ctx['order'] is not None:
            if ov_exhibition:
                setattr(ctx['order'], 'exhibition_name', ov_exhibition)
                # Backwards alias if any old template references remain
                setattr(ctx['order'], 'store_name', ov_exhibition)
            if ov_customer:
                setattr(ctx['order'], 'customer_name', ov_customer)
            if ov_badge:
                setattr(ctx['order'], 'badge_number', ov_badge)
            if ov_order_date:
                setattr(ctx['order'], 'order_date', ov_order_date)
            if ov_delivery_date:
                setattr(ctx['order'], 'delivery_date', ov_delivery_date)
    except Exception:
        pass

    # Provide provisional qr for create flow parity (same as form context key)
    if not ctx.get('order') or not getattr(ctx.get('order'), 'qr_code', None):
        try:
            import uuid as _uuid
            ctx['pre_qr_code'] = _uuid.uuid4().hex
        except Exception:
            ctx['pre_qr_code'] = ''

    # Compute a display-only serial string (English digits + uppercase) strictly for the card
    def _to_english_upper(s: str) -> str:
        """Convert Persian/Arabic-Indic digits to ASCII and uppercase letters.

        This affects only the warranty card serial display (not stored data).
        """
        out: list[str] = []
        for ch in str(s or ""):
            code = ord(ch)
            # Persian digits U+06F0..U+06F9
            if 0x06F0 <= code <= 0x06F9:
                out.append(chr(code - 0x06F0 + ord('0')))
                continue
            # Arabic-Indic digits U+0660..U+0669
            if 0x0660 <= code <= 0x0669:
                out.append(chr(code - 0x0660 + ord('0')))
                continue
            # Keep only ASCII alnum; uppercase letters; drop others for display
            ch_u = ch.upper()
            if ('A' <= ch_u <= 'Z') or ('0' <= ch_u <= '9'):
                out.append(ch_u)
            # ignore other symbols/spaces to guarantee ASCII-only output
        return ''.join(out)

    raw_code = ''
    try:
        raw_code = getattr(ctx.get('order'), 'qr_code', '') if ctx.get('order') else ''
    except Exception:
        raw_code = ''
    if not raw_code:
        raw_code = ctx.get('pre_qr_code', '')
    cleaned = ''
    try:
        cleaned = _to_english_upper(str(raw_code))
    except Exception:
        cleaned = str(raw_code or '')
    if not cleaned:
        try:
            cleaned = _to_english_upper(str(ctx.get('pre_qr_code', '')))
        except Exception:
            cleaned = str(ctx.get('pre_qr_code', '') or '')
    # Final guard: convert ASCII letters+digits to FULLWIDTH (consistent glyph set)
    def _to_fullwidth_ascii(s: str) -> str:
        out: list[str] = []
        for ch in s:
            if '0' <= ch <= '9':
                out.append(chr(ord(ch) - 48 + 0xFF10))
            elif 'A' <= ch <= 'Z':
                out.append(chr(ord(ch) - 65 + 0xFF21))
            else:
                out.append(ch)
        return ''.join(out)

    ctx['serial_display'] = _to_fullwidth_ascii((cleaned or '')[:12])
    return render(request, 'orders/warranty.html', ctx)


@login_required(login_url="/users/login/")
def warranty_card_serial(request, serial: str):
    """Render warranty card using a short serial in the path.

    This avoids putting long data (like data-URI images) in the URL. It tries
    to resolve the serial as an order's ``qr_code`` first to populate fields;
    otherwise it uses the given serial as a provisional QR code.
    """
    ctx: dict = {}

    # Try to resolve to an existing order by qr_code
    try:
        order = Order.objects.filter(qr_code=serial).first()
    except Exception:
        order = None
    if order is not None:
        ctx['order'] = order
    else:
        # No order found; carry the provided serial as a provisional QR code
        ctx['pre_qr_code'] = str(serial or '').strip()

    # Optional field overrides from query params (remain small), so print can
    # reflect unsaved edits without a long URL. These are safe, short fields.
    try:
        ov_exhibition = (request.GET.get('exhibition_name') or '').strip()
        ov_customer = (request.GET.get('customer_name') or '').strip()
        ov_badge = (request.GET.get('badge_number') or '').strip()
        ov_order_date = (request.GET.get('order_date') or '').strip()
        ov_delivery_date = (request.GET.get('delivery_date') or '').strip()
        if 'order' in ctx and ctx['order'] is not None:
            if ov_exhibition:
                setattr(ctx['order'], 'exhibition_name', ov_exhibition)
                setattr(ctx['order'], 'store_name', ov_exhibition)  # legacy alias
            if ov_customer:
                setattr(ctx['order'], 'customer_name', ov_customer)
            if ov_badge:
                setattr(ctx['order'], 'badge_number', ov_badge)
            if ov_order_date:
                setattr(ctx['order'], 'order_date', ov_order_date)
            if ov_delivery_date:
                setattr(ctx['order'], 'delivery_date', ov_delivery_date)
        else:
            # Build a lightweight context object so template fields resolve
            any_override = any([ov_exhibition, ov_customer, ov_badge, ov_order_date, ov_delivery_date])
            if any_override:
                ctx['order'] = type('OrderCtx', (), {})()
                if ov_exhibition:
                    setattr(ctx['order'], 'exhibition_name', ov_exhibition)
                    setattr(ctx['order'], 'store_name', ov_exhibition)
                if ov_customer:
                    setattr(ctx['order'], 'customer_name', ov_customer)
                if ov_badge:
                    setattr(ctx['order'], 'badge_number', ov_badge)
                if ov_order_date:
                    setattr(ctx['order'], 'order_date', ov_order_date)
                if ov_delivery_date:
                    setattr(ctx['order'], 'delivery_date', ov_delivery_date)
    except Exception:
        pass

    # Serial text used on the card (ASCII-only, then rendered as FULLWIDTH for consistency)
    def _to_english_upper(s: str) -> str:
        out: list[str] = []
        for ch in str(s or ""):
            code = ord(ch)
            if 0x06F0 <= code <= 0x06F9:  # Persian digits
                out.append(chr(code - 0x06F0 + ord('0')))
                continue
            if 0x0660 <= code <= 0x0669:  # Arabic-Indic digits
                out.append(chr(code - 0x0660 + ord('0')))
                continue
            ch_u = ch.upper()
            if ('A' <= ch_u <= 'Z') or ('0' <= ch_u <= '9'):
                out.append(ch_u)
        return ''.join(out)

    def _to_fullwidth_ascii(s: str) -> str:
        out: list[str] = []
        for ch in s:
            if '0' <= ch <= '9':
                out.append(chr(ord(ch) - 48 + 0xFF10))
            elif 'A' <= ch <= 'Z':
                out.append(chr(ord(ch) - 65 + 0xFF21))
            else:
                out.append(ch)
        return ''.join(out)

    raw_code = ''
    try:
        raw_code = getattr(ctx.get('order'), 'qr_code', '') if ctx.get('order') else ''
    except Exception:
        raw_code = ''
    if not raw_code:
        raw_code = ctx.get('pre_qr_code', '')
    cleaned = _to_english_upper(str(raw_code or ''))
    ctx['serial_display'] = _to_fullwidth_ascii((cleaned or '')[:12])

    return render(request, 'orders/warranty.html', ctx)


@login_required(login_url="/users/login/")
def order_label(request, pk: int):
    """Render the print-ready label page for a single order.

    The template shows per-product labels using the layout inspired by
    the provided Label.xlsx file.
    """

    def _to_english_upper(s: str) -> str:
        """Normalize QR code to ASCII uppercase letters and digits."""
        out: list[str] = []
        for ch in str(s or ""):
            code = ord(ch)
            # Persian digits
            if 0x06F0 <= code <= 0x06F9:
                out.append(chr(code - 0x06F0 + ord("0")))
                continue
            # Arabic-Indic digits
            if 0x0660 <= code <= 0x0669:
                out.append(chr(code - 0x0660 + ord("0")))
                continue
            ch_u = ch.upper()
            if ("A" <= ch_u <= "Z") or ("0" <= ch_u <= "9"):
                out.append(ch_u)
        return "".join(out)

    def _to_fullwidth_ascii(s: str) -> str:
        """Convert ASCII letters/digits to FULLWIDTH glyphs for visual consistency."""
        out: list[str] = []
        for ch in s:
            if "0" <= ch <= "9":
                out.append(chr(ord(ch) - 48 + 0xFF10))
            elif "A" <= ch <= "Z":
                out.append(chr(ord(ch) - 65 + 0xFF21))
            else:
                out.append(ch)
        return "".join(out)

    order = get_object_or_404(
        Order.objects.prefetch_related("items__product__product_model"),
        pk=pk,
    )

    # Build the same base serial string used on the warranty card, limited to 12 chars.
    raw_code = getattr(order, "qr_code", "") or ""
    cleaned = _to_english_upper(raw_code)
    serial_base = _to_fullwidth_ascii((cleaned or "")[:12]) if cleaned else ""

    # Build per-item product codes: "<serial_base> <index>" where the whole
    # string is rendered using the same FULLWIDTH ASCII mapping so that both
    # the base and the appended index are visually English (non-localized).
    label_items = []
    items = list(order.items.all())
    if cleaned:
        for idx, item in enumerate(items, start=1):
            ascii_code = f"{(cleaned or '')[:12]} {idx}"
            fullwidth_code = _to_fullwidth_ascii(ascii_code)
            label_items.append({"item": item, "code": fullwidth_code})
    else:
        for item in items:
            label_items.append({"item": item, "code": ""})

    context = {
        "order": order,
        "serial_base": serial_base,
        "label_items": label_items,
    }
    return render(request, "orders/label.html", context)


def qr_image_svg(request, code: str):
    """Generate and return a real QR code as SVG for the given ``code``.

    - Uses the lightweight ``qrcode`` library with the SVG image factory,
      so no Pillow dependency is required.
    - Falls back to a simple placeholder SVG only if the library is missing
      or generation fails for any reason.
    """
    try:
        import qrcode  # type: ignore
        from qrcode.image.svg import SvgPathImage  # type: ignore

        qr = qrcode.QRCode(
            version=None,  # automatically determine the minimal version
            error_correction=qrcode.constants.ERROR_CORRECT_M,
            box_size=8,  # cell size; overall SVG will scale without blur
            border=0,    # outer quiet zone handled by layout CSS
        )
        qr_data = code or ""
        try:
            if code:
                if not (code.startswith("http://") or code.startswith("https://")):
                    public_url = request.build_absolute_uri(
                        reverse("orders:public_order_summary", args=[code])
                    )
                    qr_data = public_url
                else:
                    qr_data = code
        except Exception:
            qr_data = code or ""

        qr.add_data(qr_data)
        qr.make(fit=True)
        img = qr.make_image(image_factory=SvgPathImage)
        # ``to_string()`` returns bytes containing the SVG markup
        svg_bytes = img.to_string()
        return HttpResponse(svg_bytes, content_type="image/svg+xml; charset=utf-8")
    except Exception:
        # Fallback placeholder to ensure something is visible and printable
        svg = f'''<?xml version="1.0" encoding="UTF-8"?>
<svg xmlns="http://www.w3.org/2000/svg" width="280" height="280" viewBox="0 0 280 280">
  <rect x="0" y="0" width="280" height="280" fill="#fff" stroke="#000"/>
  <rect x="10" y="10" width="260" height="260" fill="#fff" stroke="#000" stroke-width="2"/>
  <text x="50%" y="50%" dominant-baseline="middle" text-anchor="middle" font-family="monospace" font-size="12">{code}</text>
  <text x="50%" y="88%" dominant-baseline="middle" text-anchor="middle" font-family="sans-serif" font-size="10">QR placeholder</text>
</svg>'''
        return HttpResponse(svg, content_type='image/svg+xml; charset=utf-8')


def public_order_summary(request, serial: str):
    """Public read-only order summary page for QR scans.

    This view is intentionally unauthenticated so that customers can scan
    a QR code from any device and see the current status of their order,
    related jobs, process flow and daily production logs.
    """
    serial = (serial or "").strip()
    order = (
        Order.objects.filter(qr_code=serial)
        .prefetch_related("items__product__product_model", "jobs__product__product_model")
        .first()
    )
    if not order:
        return render(
            request,
            "orders/public_order_summary.html",
            {
                "order": None,
                "serial": serial,
                "jobs": [],
                "overall_progress": 0,
            },
            status=404,
        )

    jobs_qs = (
        ProductionJob.objects.filter(order=order)
        .select_related("product", "part")
        .order_by("created_at")
    )

    jobs_data = []
    progress_values: list[int] = []

    for job in jobs_qs:
        allowed_sections = getattr(job, "allowed_sections", []) or []
        progress_payload = _build_progress_state(job, allowed_sections)
        flow_items = progress_payload.get("items", [])
        cursor = int(progress_payload.get("cursor") or 0)
        flow_length = int(progress_payload.get("flow_length") or 0)

        if flow_length > 0:
            percent = int(round((cursor / flow_length) * 100))
        else:
            percent = 0
        # If job is marked completed or has finished_at, clamp to 100%.
        if getattr(job, "status", "") == "completed" or getattr(job, "finished_at", None):
            percent = 100

        progress_values.append(percent)

        logs = (
            ProductionLog.objects.filter(job=job)
            .select_related("product", "part", "user")
            .order_by("logged_at")
        )

        jobs_data.append(
            {
                "instance": job,
                "progress_percent": percent,
                "flow_items": flow_items,
                "logs": logs,
            }
        )

    overall_progress = 0
    if progress_values:
        overall_progress = int(round(sum(progress_values) / len(progress_values)))

    # Reuse the same QR serial display logic as the warranty card.
    def _to_english_upper(s: str) -> str:
        out: list[str] = []
        for ch in str(s or ""):
            code = ord(ch)
            if 0x06F0 <= code <= 0x06F9:
                out.append(chr(code - 0x06F0 + ord("0")))
                continue
            if 0x0660 <= code <= 0x0669:
                out.append(chr(code - 0x0660 + ord("0")))
                continue
            ch_u = ch.upper()
            if ("A" <= ch_u <= "Z") or ("0" <= ch_u <= "9"):
                out.append(ch_u)
        return "".join(out)

    def _to_fullwidth_ascii(s: str) -> str:
        out: list[str] = []
        for ch in s:
            if "0" <= ch <= "9":
                out.append(chr(ord(ch) - 48 + 0xFF10))
            elif "A" <= ch <= "Z":
                out.append(chr(ord(ch) - 65 + 0xFF21))
            else:
                out.append(ch)
        return "".join(out)

    raw_code = getattr(order, "qr_code", "") or serial
    cleaned = _to_english_upper(raw_code)
    serial_display = _to_fullwidth_ascii((cleaned or "")[:12]) if cleaned else ""

    context = {
        "order": order,
        "serial": serial,
        "serial_display": serial_display,
        "jobs": jobs_data,
        "overall_progress": overall_progress,
    }
    return render(request, "orders/public_order_summary.html", context)

class _ItemsSaverMixin:
    """Helpers to persist requested_products -> OrderItem rows."""

    def _save_items_from_requested(self, order, requested_map, old_status=None):
        """
        Persist the requested products into OrderItem rows and adjust the parts inventory.

        Parameters:
            order:        The Order instance being created or updated.
            requested_map: A dict mapping product_id to quantity.
            old_status:   The previous status of the order before this save, or None for new orders.

        The logic follows these rules:
            - If old_status is not None and not equal to 'لغو شده', the existing items' quantities
              were previously deducted from inventory and must be added back before applying new items.
            - If the new order status is not 'لغو شده', the new items' quantities will be deducted
              from inventory. Otherwise, inventory is left unchanged for cancelled orders.
            - Inventory adjustments apply to the parts' ``stock_cut`` and ``stock_cnc_tools`` fields (not ``stock_assembly``).
            - Matching of parts is case-insensitive on the part name and its associated product model name.
        """
        # Normalize requested_map keys to ints
        pids = [int(pid) for pid in requested_map.keys()]
        products = {p.id: p for p in Product.objects.filter(id__in=pids)}

        new_status = getattr(order, "status", None)
        previous_status = old_status

        # Determine whether to reverse inventory for existing items
        reverse_inventory = False
        if previous_status is not None and str(previous_status).strip() != 'لغو شده':
            reverse_inventory = True

        if reverse_inventory:
            # Add back quantities for existing items
            for existing in order.items.all():
                comps = getattr(existing.product, 'components', []) or []
                for comp in comps:
                    part_name = comp.get('part_name') or comp.get('name') or comp.get('part')
                    if not part_name:
                        continue
                    qty_required = int(comp.get('qty') or 1)
                    delta = qty_required * existing.quantity


        # Remove all existing items (they will be recreated)
        try:
            order.items.all().delete()
        except Exception:
            OrderItem.objects.filter(order=order).delete()

        # Build new OrderItem instances
        bulk_items = []
        for pid, qty in requested_map.items():
            product = products.get(int(pid))
            q = int(qty) if qty is not None else 1
            if product and q > 0:
                bulk_items.append(OrderItem(order=order, product=product, quantity=q))

        # Save new items to DB
        if bulk_items:
            OrderItem.objects.bulk_create(bulk_items)

            # Decide whether to deduct inventory for new items
            deduct_inventory = True
            if str(new_status).strip() == 'لغو شده':
                deduct_inventory = False

            if deduct_inventory:
                for item in bulk_items:
                    comps = getattr(item.product, 'components', []) or []
                    for comp in comps:
                        part_name = comp.get('part_name') or comp.get('name') or comp.get('part')
                        if not part_name:
                            continue
                        qty_required = int(comp.get('qty') or 1)
                        delta = qty_required * item.quantity


class OrderCreateView(LoginRequiredMixin, _ItemsSaverMixin, CreateView):
    login_url = "/users/login/"
    model = Order
    form_class = OrderForm
    template_name = 'orders/orders_form.html'
    success_url = reverse_lazy('orders:list')

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        # Populate product model choices for the checkbox filters.  Only the

        # ``product_type`` list has been removed.
        try:
            from inventory.models import ProductModel
            ctx["model_choices"] = [(m.name, m.name) for m in ProductModel.objects.all().order_by('name')]
        except Exception:
            ctx["model_choices"] = []
        ctx["is_editing"] = False

        # Generate a provisional QR code for the new order on initial GET so the
        # form can display a unique QR immediately. Persist it via a hidden
        # input; on submit we will save it as the order's qr_code.
        form = ctx.get('form')
        try:
            if form is not None and not form.is_bound:
                ctx['pre_qr_code'] = _ensure_unique_qr_code()
            elif form is not None and form.is_bound:
                ctx['pre_qr_code'] = (form.data.get('qr_code') or '').strip()
        except Exception:
            ctx['pre_qr_code'] = ''

        selected_models = []
        selected_product_ids = []
        if form is not None:
            if form.is_bound:
                selected_models = [m.strip() for m in form.data.getlist('product_models') if m.strip()]
                selected_product_ids = _extract_product_ids(form.data.get('requested_products'))
            else:
                initial_models = form.initial.get('product_models', []) if isinstance(form.initial, dict) else []
                selected_models = [m.strip() for m in initial_models if isinstance(m, str) and m.strip()]
                initial_requested = None
                if isinstance(form.initial, dict):
                    initial_requested = form.initial.get('requested_products')
                selected_product_ids = _extract_product_ids(initial_requested)

        # Available production jobs for selection.  Jobs that have not yet

        # includes the primary key, job number, the related product's ID and
        # name, the Persian display of its label (tag), and a Tailwind CSS
        # class for coloring.  These details are used in the template to
        # render a checklist where users can select job numbers.
        try:

            # production_line module.  The ProductionJob model was moved
            # into ``jobs.models``; importing from there avoids relying on
            # backwards-compatibility shims.
            from jobs.models import ProductionJob
            jobs_qs = ProductionJob.objects.select_related('product', 'product__product_model').filter(order__isnull=True)
            if selected_models:
                jobs_qs = jobs_qs.filter(product__product_model__name__in=selected_models)
            if selected_product_ids:
                jobs_qs = jobs_qs.filter(product_id__in=selected_product_ids)
            else:
                jobs_qs = jobs_qs.none()
            job_options = []
            for job in jobs_qs:
                # Determine product id and name; if missing set None/''
                pid = getattr(job.product, 'id', None)
                pname = getattr(job.product, 'name', '') if getattr(job, 'product', None) else ''
                pmodel = ''
                try:
                    pmodel = job.product.product_model.name if job.product and job.product.product_model else ''
                except Exception:
                    pmodel = ''
                # Determine Persian label and corresponding color class
                label_display = job.get_job_label_display() if hasattr(job, 'get_job_label_display') else job.job_label
                # Map job_label to a background color class similar to reports/jobs.html
                match = {
                    'in_progress': ('bg-gray-500 text-white', 'در حال ساخت'),
                    'completed':   ('bg-green-400 text-white', 'تولید شده'),
                    'scrapped':    ('bg-red-600 text-white', 'اسقاط'),
                    'warranty':    ('bg-yellow-300 text-black', 'گارانتی'),
                    'repaired':    ('bg-blue-600 text-white', 'تعمیرات'),
                    'deposit':     ('bg-yellow-600 text-white', 'امانی'),
                }
                color_cls, persian_label = match.get(job.job_label, ('bg-gray-400 text-white', label_display))
                job_options.append({
                    'id': job.pk,
                    'job_number': job.job_number,
                    'product_id': pid,
                    'product_name': pname,
                    'product_model': pmodel,
                    'label': persian_label,
                    'color_class': color_cls,
                })
            ctx['available_jobs'] = job_options
        except Exception:
            ctx['available_jobs'] = []
        if form is not None:
            if form.is_bound:
                try:
                    ctx['selected_job_ids'] = [int(val) for val in form.data.getlist('job_numbers')]
                except Exception:
                    ctx['selected_job_ids'] = list(form.data.getlist('job_numbers'))
            else:
                ctx['selected_job_ids'] = []
        else:
            ctx['selected_job_ids'] = []  # no preselected jobs for create
        return ctx

    @transaction.atomic
    def form_valid(self, form):
        # Save with a stable QR code if provided from the form
        qr_from_form = (self.request.POST.get('qr_code') or '').strip()
        obj = form.save(commit=False)
        if not getattr(obj, 'qr_code', None) and qr_from_form:
            obj.qr_code = _ensure_unique_qr_code(qr_from_form)
        self.object = obj
        try:
            self.object.save()
        except IntegrityError as exc:
            message = str(exc)
            if 'orders_order_qr_code_key' in message:
                self.object.qr_code = _generate_unique_qr_code()
                self.object.save()
            elif 'badge_number' in message or 'orders_order_badge_number_key' in message:
                form.add_error('badge_number', "شماره بیجک وارد شده تکراری است. لطفاً شماره دیگری وارد کنید.")
                return self.form_invalid(form)
            else:
                raise
        # Ensure a unique QR code is assigned on creation if missing.
        try:
            if not getattr(self.object, 'qr_code', None):
                self.object.qr_code = _generate_unique_qr_code()
                self.object.save(update_fields=["qr_code"])
        except Exception:
            # If QR code assignment fails, proceed; backfill command can populate later.
            pass
        requested = form.cleaned_data.get("requested_products") or {}
        self._save_items_from_requested(self.object, requested, old_status=None)

        # Assign selected production jobs (if any) to this order.  Jobs are
        # linked both to the order and the corresponding OrderItem based on
        # the product.  Jobs selected via the form are identified by their
        # primary key.  Any job that is already assigned to another order is
        # ignored; the form validation prevents this situation from occurring.
        selected_job_ids = form.cleaned_data.get('job_numbers') or []
        if selected_job_ids:
            try:
                from jobs.models import ProductionJob
                jobs_to_assign = ProductionJob.objects.select_related('product').filter(pk__in=[int(i) for i in selected_job_ids])
                for job in jobs_to_assign:
                    # Skip if the job is already assigned to this order
                    if job.order_id == self.object.id:
                        continue
                    # Assign to this order and matching OrderItem
                    job.order = self.object
                    # Match by product id; there could be multiple quantities but only one
                    # OrderItem per product.  We assign the job to that item.
                    order_item = None
                    try:
                        if job.product_id:
                            order_item = self.object.items.filter(product_id=job.product_id).first()
                    except Exception:
                        order_item = None
                    job.order_item = order_item
                    job.save(update_fields=['order', 'order_item'])
            except Exception:
                pass

        # Assign a unique job number to each created OrderItem if none exists
        # already.  Each item receives its own tracking code irrespective of
        # the external production jobs selected above.
        import uuid
        for item in self.object.items.all():
            if not getattr(item, 'job_number', None):
                item.job_number = f"JOB-{uuid.uuid4().hex[:10].upper()}"
                item.save(update_fields=["job_number"])

        messages.success(self.request, "سفارش ثبت شد.")
        return redirect(self.success_url)

class OrderUpdateView(LoginRequiredMixin, _ItemsSaverMixin, UpdateView):
    login_url = "/users/login/"
    model = Order
    form_class = OrderForm
    template_name = 'orders/orders_form.html'
    success_url = reverse_lazy('orders:list')

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        # Provide the same model choices for edit mode.  See OrderCreateView.
        try:
            from inventory.models import ProductModel
            ctx["model_choices"] = [(m.name, m.name) for m in ProductModel.objects.all().order_by('name')]
        except Exception:
            ctx["model_choices"] = []
        ctx["is_editing"] = True
        # Provide QR code value for display on edit
        try:
            obj = getattr(self, 'object', None)
            ctx['pre_qr_code'] = getattr(obj, 'qr_code', '') if obj else ''
        except Exception:
            ctx['pre_qr_code'] = ''

        form = ctx.get('form')
        selected_models = []
        selected_product_ids = []
        if form is not None:
            if form.is_bound:
                selected_models = [m.strip() for m in form.data.getlist('product_models') if m.strip()]
                selected_product_ids = _extract_product_ids(form.data.get('requested_products'))
            else:
                initial_models = form.initial.get('product_models', []) if isinstance(form.initial, dict) else []
                selected_models = [m.strip() for m in initial_models if isinstance(m, str) and m.strip()]
                initial_requested = None
                if isinstance(form.initial, dict):
                    initial_requested = form.initial.get('requested_products')
                elif self.object:
                    try:
                        items_mapping = {it.product_id: int(it.quantity or 1) for it in self.object.items.all()}
                        initial_requested = items_mapping
                    except Exception:
                        initial_requested = None
                selected_product_ids = _extract_product_ids(initial_requested)

        # Provide available jobs and preselected jobs.  On edit, include
        # unassigned jobs plus jobs already linked to this order.
        try:
            from jobs.models import ProductionJob
            obj = self.object
            jobs_qs = ProductionJob.objects.select_related('product', 'product__product_model').filter(
                models.Q(order__isnull=True) | models.Q(order=obj)
            )
            if selected_models:
                jobs_qs = jobs_qs.filter(
                    models.Q(product__product_model__name__in=selected_models) | models.Q(order=obj)
                )
            else:
                jobs_qs = jobs_qs.filter(order=obj)
            if selected_product_ids:
                jobs_qs = jobs_qs.filter(
                    models.Q(product_id__in=selected_product_ids) | models.Q(order=obj)
                )
            else:
                jobs_qs = jobs_qs.filter(order=obj)
            job_options = []
            for job in jobs_qs:
                pid = getattr(job.product, 'id', None)
                pname = getattr(job.product, 'name', '') if getattr(job, 'product', None) else ''
                pmodel = ''
                try:
                    pmodel = job.product.product_model.name if job.product and job.product.product_model else ''
                except Exception:
                    pmodel = ''
                label_display = job.get_job_label_display() if hasattr(job, 'get_job_label_display') else job.job_label
                match = {
                    'in_progress': ('bg-gray-500 text-white', 'در حال ساخت'),
                    'completed':   ('bg-green-400 text-white', 'تولید شده'),
                    'scrapped':    ('bg-red-600 text-white', 'اسقاط'),
                    'warranty':    ('bg-yellow-300 text-black', 'گارانتی'),
                    'repaired':    ('bg-blue-600 text-white', 'تعمیرات'),
                    'deposit':     ('bg-yellow-600 text-white', 'امانی'),
                }
                color_cls, persian_label = match.get(job.job_label, ('bg-gray-400 text-white', label_display))
                job_options.append({
                    'id': job.pk,
                    'job_number': job.job_number,
                    'product_id': pid,
                    'product_name': pname,
                    'product_model': pmodel,
                    'label': persian_label,
                    'color_class': color_cls,
                })
            ctx['available_jobs'] = job_options
            # Preselect jobs already associated with this order
            selected = ProductionJob.objects.filter(order=obj).values_list('id', flat=True)
            preselected = list(selected)
            if form is not None and form.is_bound:
                try:
                    preselected = [int(val) for val in form.data.getlist('job_numbers')]
                except Exception:
                    preselected = list(form.data.getlist('job_numbers'))
            ctx['selected_job_ids'] = preselected
        except Exception:
            ctx['available_jobs'] = []
            ctx['selected_job_ids'] = []
        return ctx

    @transaction.atomic
    def form_valid(self, form):

        try:
            prev_order = self.get_object()
            old_status = getattr(prev_order, "status", None)
        except Exception:
            old_status = None
        # Respect QR code provided by the form if still missing
        qr_from_form = (self.request.POST.get('qr_code') or '').strip()
        obj = form.save(commit=False)
        if not getattr(obj, 'qr_code', None) and qr_from_form:
            obj.qr_code = qr_from_form
        self.object = obj
        try:
            self.object.save()
        except IntegrityError as exc:
            message = str(exc)
            if 'badge_number' in message or 'orders_order_badge_number_key' in message:
                form.add_error('badge_number', "شماره بیجک وارد شده تکراری است. لطفاً شماره دیگری وارد کنید.")
                return self.form_invalid(form)
            raise

        try:
            import uuid as _uuid
            if not getattr(self.object, 'qr_code', None):
                self.object.qr_code = _uuid.uuid4().hex
                self.object.save(update_fields=["qr_code"])
        except Exception:
            pass
        requested = form.cleaned_data.get("requested_products") or {}
        self._save_items_from_requested(self.object, requested, old_status=old_status)


        # associated with this order that are no longer selected.  Then
        # assign any newly selected jobs.  Jobs selected via the form are
        # identified by their primary keys.  Validation on the form ensures
        # jobs are not concurrently assigned to another order.
        selected_job_ids = form.cleaned_data.get('job_numbers') or []
        try:
            from jobs.models import ProductionJob
            current_jobs_qs = ProductionJob.objects.filter(order=self.object)
            selected_set = {int(i) for i in selected_job_ids}
            # Detach jobs not in the selected set
            for job in current_jobs_qs:
                if job.pk not in selected_set:
                    job.order = None
                    job.order_item = None
                    job.save(update_fields=['order', 'order_item'])
            # Assign newly selected jobs
            to_assign_ids = selected_set - set(current_jobs_qs.values_list('id', flat=True))
            if to_assign_ids:
                jobs_to_assign = ProductionJob.objects.select_related('product').filter(pk__in=to_assign_ids)
                for job in jobs_to_assign:
                    job.order = self.object
                    # Associate with matching OrderItem by product
                    order_item = None
                    try:
                        if job.product_id:
                            order_item = self.object.items.filter(product_id=job.product_id).first()
                    except Exception:
                        order_item = None
                    job.order_item = order_item
                    job.save(update_fields=['order', 'order_item'])
        except Exception:
            pass

        # Ensure all items have a job number assigned as a fallback.  This
        # covers cases where no external jobs were selected for a product.
        import uuid
        for item in self.object.items.all():
            if not getattr(item, 'job_number', None):
                item.job_number = f"JOB-{uuid.uuid4().hex[:10].upper()}"
                item.save(update_fields=["job_number"])

        messages.success(self.request, "سفارش به‌روزرسانی شد.")
        return redirect(self.success_url)

class OrderBulkDeleteView(LoginRequiredMixin, View):
    login_url = "/users/login/"
    def post(self, request, *args, **kwargs):
        ids = request.POST.getlist('selected_orders')
        if ids:
            orders_to_delete = Order.objects.filter(id__in=ids).prefetch_related('items', 'items__product')
            deleted_count = 0
            for order in orders_to_delete:

                if str(order.status).strip() != 'لغو شده':
                    for item in order.items.all():
                        comps = getattr(item.product, 'components', []) or []
                        for comp in comps:
                            part_name = comp.get('part_name') or comp.get('name') or comp.get('part')
                            if not part_name:
                                continue
                            qty_required = int(comp.get('qty') or 1)
                            delta = qty_required * item.quantity

                order.delete()
                deleted_count += 1
            messages.success(request, f"{deleted_count} سفارش حذف شد.")
        else:
            messages.warning(request, "هیچ سفارشی انتخاب نشده است.")
        return redirect('orders:list')

class ProductsByModelsView(LoginRequiredMixin, View):
    login_url = "/users/login/"
    """
    Return products filtered by one or more product model names.

    Clients should supply the list of desired model names via the
    ``models`` or ``models[]`` query parameters.  The response
    contains a list of products, each with its primary key (``id``),
    its name, and its associated model name under the ``model`` key.
    """

    def get(self, request, *args, **kwargs):
        models_sel = request.GET.getlist('models[]') or request.GET.getlist('models')
        if not models_sel:
            return JsonResponse({'products': []})
        # Filter products by the names of their related product models and
        # order by model then product name.
        qs = Product.objects.filter(product_model__name__in=models_sel).select_related('product_model').order_by('product_model__name', 'name')
        data = [
            {
                'id': p.id,
                'name': p.name,
                'model': p.product_model.name if p.product_model else ''
            }
            for p in qs
        ]
        return JsonResponse({'products': data})


class JobsBySelectionView(LoginRequiredMixin, View):
    login_url = "/users/login/"
    """
    Return available production jobs filtered by selected models and products.

    Query params:
    - models[]: list of product model names (optional)
    - product_ids[]: list of product IDs (required to show jobs on create)
    - order_id: optional current order id (to include jobs already linked on edit)

    Response JSON:
    { "jobs": [ { id, job_number, product_id, product_name, product_model, label, color_class } ] }
    """

    def get(self, request, *args, **kwargs):
        models_sel = request.GET.getlist('models[]') or request.GET.getlist('models')
        product_ids_raw = request.GET.getlist('product_ids[]') or request.GET.getlist('product_ids')
        try:
            order_id = int(request.GET.get('order_id')) if request.GET.get('order_id') else None
        except Exception:
            order_id = None

        # Normalize IDs list
        product_ids: list[int] = []
        for v in product_ids_raw:
            try:
                product_ids.append(int(v))
            except Exception:
                continue

        # Import the ProductionJob model from jobs app
        try:
            from jobs.models import ProductionJob  # type: ignore
        except Exception:
            return JsonResponse({"jobs": []})

        # Base queryset: unassigned jobs. On edit, include jobs linked to this order as well.
        base_qs = ProductionJob.objects.select_related('product', 'product__product_model')
        if order_id:
            base_qs = base_qs.filter(models.Q(order__isnull=True) | models.Q(order_id=order_id))
        else:
            base_qs = base_qs.filter(order__isnull=True)

        # Apply model filter if provided
        if models_sel:
            base_qs = base_qs.filter(product__product_model__name__in=models_sel)

        # Apply product filter; if not provided, on create we return empty to avoid noise
        if product_ids:
            qs = base_qs.filter(product_id__in=product_ids)
        else:
            # If editing an order, we still may want to include its already-linked jobs
            if order_id:
                qs = base_qs.filter(order_id=order_id)
            else:
                qs = base_qs.none()

        # Prepare response payload
        jobs_data = []
        # Mapping of job_label to color classes and Persian label used elsewhere in the app
        label_map = {
            'in_progress': ('bg-gray-500 text-white', 'در حال ساخت'),
            'completed':   ('bg-green-400 text-white', 'تولید شده'),
            'scrapped':    ('bg-red-600 text-white', 'اسقاط'),
            'warranty':    ('bg-yellow-300 text-black', 'گارانتی'),
            'repaired':    ('bg-blue-600 text-white', 'تعمیرات'),
            'deposit':     ('bg-yellow-600 text-white', 'امانی'),
        }
        for job in qs:
            try:
                pid = getattr(job, 'product_id', None)
                pname = getattr(job.product, 'name', '') if getattr(job, 'product', None) else ''
                try:
                    pmodel = job.product.product_model.name if job.product and job.product.product_model else ''
                except Exception:
                    pmodel = ''
                label_display = job.get_job_label_display() if hasattr(job, 'get_job_label_display') else getattr(job, 'job_label', '')
                color_cls, fa_label = label_map.get(getattr(job, 'job_label', ''), ('bg-gray-400 text-white', label_display))
                jobs_data.append({
                    'id': job.pk,
                    'job_number': job.job_number,
                    'product_id': pid,
                    'product_name': pname,
                    'product_model': pmodel,
                    'label': fa_label,
                    'color_class': color_cls,
                })
            except Exception:
                # Skip malformed jobs safely
                continue

        return JsonResponse({"jobs": jobs_data})


class OrderStageUpdateView(LoginRequiredMixin, View):
    login_url = "/users/login/"
    """Allow inline toggling of the production stage from the list view."""

    def post(self, request, pk, *args, **kwargs):
        order = get_object_or_404(Order, pk=pk)
        data = {}
        if request.body:
            try:
                data = json.loads(request.body.decode('utf-8'))
            except (ValueError, TypeError, AttributeError):
                data = {}
        stage_value = (data.get('stage') or request.POST.get('stage') or '').strip()
        valid_stages = [choice[0] for choice in Order.STAGE_CHOICES]
        if stage_value not in valid_stages:
            return JsonResponse({'success': False, 'error': 'invalid_stage'}, status=400)
        order.current_stage = stage_value
        order.save(update_fields=['current_stage'])
        return JsonResponse({'success': True, 'stage': stage_value})


class LiveOrdersFeedView(LoginRequiredMixin, View):
    login_url = "/users/login/"
    """Return a lightweight JSON feed of the latest orders for dashboard sync."""

    def get(self, request, *args, **kwargs):
        try:
            limit = int(request.GET.get('limit', 200))
        except (TypeError, ValueError):
            limit = 200
        limit = max(1, min(limit, 500))

        search = (request.GET.get('q') or '').strip()
        qs = Order.objects.order_by('-id')
        if search:
            qs = qs.filter(
                models.Q(badge_number__icontains=search) |
                models.Q(subscription_code__icontains=search) |
                models.Q(customer_name__icontains=search) |
                models.Q(exhibition_name__icontains=search)
            )

        data = []
        for order in qs[:limit]:
            data.append({
                'id': order.id,
                'badge_number': order.badge_number or '',
                'subscription_code': order.subscription_code or '',
                'customer_name': order.customer_name or '',
                'model': order.model or '',
                'status': order.status,
                'status_display': order.get_status_display(),
            })

        return JsonResponse({'orders': data})

