from __future__ import annotations

from decimal import Decimal
from io import BytesIO
from typing import Iterable, List, Sequence

from django.http import HttpResponse
from django.utils import timezone
import jdatetime

from openpyxl import Workbook
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


def sanitize_value(raw: object) -> object:
    """
    Prepare a value for XLSX cells while preserving numeric types.

    - Keep ints/floats/Decimals numeric so Excel treats them as numbers.
    - Strip illegal control chars from text.
    - Booleans render as Persian yes/no for readability.
    """
    if raw is None:
        return ""
    if isinstance(raw, bool):
        return "بله" if raw else "خیر"
    if isinstance(raw, int):
        return raw
    if isinstance(raw, float):
        # Preserve integers without a trailing .0 where possible
        return int(raw) if raw.is_integer() else raw
    if isinstance(raw, Decimal):
        try:
            as_int = int(raw)
            if raw == Decimal(as_int):
                return as_int
        except Exception:
            pass
        try:
            return float(raw)
        except Exception:
            return str(raw)
    text = str(raw)
    try:
        text = ILLEGAL_CHARACTERS_RE.sub("", text)
    except Exception:
        pass
    return text


def base_styles():
    """Return the shared style objects used across XLSX exports."""
    title_font = Font(name="Tahoma", bold=True, size=14)
    header_font = Font(name="Tahoma", bold=True, size=11)
    cell_font = Font(name="Tahoma", size=11)
    center_header = Alignment(horizontal="center", vertical="center", wrap_text=True)
    right_cell = Alignment(horizontal="right", vertical="center", wrap_text=True)
    header_fill = PatternFill("solid", fgColor="FFF9FAFB")
    thin_side = Side(style="thin", color="FFE5E7EB")
    border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    center_cell = Alignment(horizontal="center", vertical="center", wrap_text=True)
    return {
        "title_font": title_font,
        "header_font": header_font,
        "cell_font": cell_font,
        "center_header": center_header,
        "right_cell": right_cell,
        "center_cell": center_cell,
        "header_fill": header_fill,
        "border": border,
    }


def _safe_table_name(base: str, existing: set[str]) -> str:
    cleaned = "".join(ch if ch.isalnum() else "_" for ch in (base or "Table"))
    if not cleaned:
        cleaned = "Table"
    if cleaned[0].isdigit():
        cleaned = f"T{cleaned}"
    candidate = cleaned
    counter = 1
    while candidate in existing:
        candidate = f"{cleaned}_{counter}"
        counter += 1
    return candidate


def write_table(
    ws,
    headers: Sequence[str],
    rows: Iterable[Sequence[object]],
    *,
    start_row: int = 1,
    column_widths: Sequence[int] | None = None,
    table_name: str | None = None,
    add_table: bool = True,
):
    """Write a styled table (header + rows) and optionally add an Excel table style."""
    styles = base_styles()
    header_row_idx = start_row

    def is_name_desc(label: str) -> bool:
        """Identify name/description columns to keep RTL alignment."""
        needle = str(label)
        return ("نام" in needle) or ("توضیح" in needle) or ("description" in needle.lower()) or ("name" in needle.lower())

    # Header
    for col_idx, label in enumerate(headers, start=1):
        c = ws.cell(row=header_row_idx, column=col_idx, value=label)
        c.font = styles["header_font"]
        c.alignment = styles["center_header"]
        c.fill = styles["header_fill"]
        c.border = styles["border"]

    # Data rows
    row_idx = header_row_idx + 1
    for data_row in rows:
        for col_idx, raw_value in enumerate(data_row, start=1):
            value = sanitize_value(raw_value)
            c = ws.cell(row=row_idx, column=col_idx, value=value)
            c.font = styles["cell_font"]
            # Center all cells except name/description columns
            if is_name_desc(headers[col_idx - 1]):
                c.alignment = styles["right_cell"]
            else:
                c.alignment = styles["center_cell"]
            c.border = styles["border"]
        row_idx += 1

    # Column widths
    widths = column_widths or []
    default_width = 24
    for col_idx in range(1, len(headers) + 1):
        letter = get_column_letter(col_idx)
        width = widths[col_idx - 1] if col_idx - 1 < len(widths) else default_width
        try:
            ws.column_dimensions[letter].width = width
        except Exception:
            pass

    data_end = row_idx - 1
    if add_table and data_end >= header_row_idx:
        # openpyxl stores tables differently across versions (_tables may be dict or list)
        raw_tables = getattr(ws, "_tables", [])  # type: ignore[attr-defined]
        if isinstance(raw_tables, dict):
            existing = set(raw_tables.keys())
        else:
            existing = {
                getattr(t, "displayName", str(t))  # type: ignore[attr-defined]
                for t in raw_tables
            }
        name = table_name or f"Table{len(existing) + 1}"
        name = _safe_table_name(name, existing)
        ref = f"A{header_row_idx}:{get_column_letter(len(headers))}{data_end}"
        table = Table(displayName=name, ref=ref)
        style = TableStyleInfo(
            name="TableStyleMedium9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        table.tableStyleInfo = style
        ws.add_table(table)

    return header_row_idx, data_end


def build_table_response(
    *,
    sheet_title: str,
    report_title: str | None,
    headers: List[str],
    rows: Iterable[Sequence[object]],
    filename: str,
    column_widths: Sequence[int] | None = None,
    subtitle: str | None = None,
    include_timestamp: bool = True,
    table_name: str | None = None,
    right_to_left: bool = True,
):
    """
    Build a single-sheet XLSX response with a styled data table.

    Adds a merged title row, optional timestamp subtitle and a banded Excel table
    so the exported sheet looks like a proper grid.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title or "گزارش"
    if right_to_left:
        try:
            ws.sheet_view.rightToLeft = True
        except Exception:
            pass

    styles = base_styles()
    row_idx = 1

    if report_title:
        ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=len(headers))
        c = ws.cell(row=row_idx, column=1, value=report_title)
        c.font = styles["title_font"]
        c.alignment = styles["center_header"]
        row_idx += 1

    ts_text = subtitle
    if include_timestamp and not subtitle:
        try:
            gnow = timezone.localtime(timezone.now())
            ts_text = jdatetime.datetime.fromgregorian(datetime=gnow).strftime("%Y/%m/%d %H:%M")
        except Exception:
            ts_text = None
    if ts_text:
        ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=len(headers))
        c = ws.cell(row=row_idx, column=1, value=f"تاریخ تهیه: {ts_text}")
        c.font = styles["cell_font"]
        c.alignment = styles["right_cell"]
        row_idx += 1

    write_table(
        ws,
        headers=headers,
        rows=rows,
        start_row=row_idx,
        column_widths=column_widths,
        table_name=table_name,
        add_table=True,
    )

    bio = BytesIO()
    wb.save(bio)
    resp = HttpResponse(
        bio.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = f"attachment; filename={filename}"
    return resp
